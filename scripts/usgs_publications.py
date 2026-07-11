#!/usr/bin/env python3
"""Normalize the frozen USGS rare-earth publications used by the V3 site.

This module deliberately keeps two views of the 2026 Mineral Commodity
Summaries (MCS) data.  ``raw_*`` fields reproduce the ScienceBase CSV, while
``current_*`` and parsed value fields apply the four changes documented in the
official MCS 2026 revision history and visible in the current version 1.3 PDF.
No raw value or note is overwritten.

The 2022 Minerals Yearbook (MYB) parser is intentionally limited to table T8,
world mine production.  The remaining MYB tables are preserved in the frozen
workbook but are not mixed into the site's DataWeb partner-trade record.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import warnings
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
RETRIEVED_AT = "2026-07-10"
SOURCE_HUB_URL = (
    "https://www.usgs.gov/centers/national-minerals-information-center/"
    "rare-earths-statistics-and-information"
)
MCS_DATA_RELEASE_URL = (
    "https://data.usgs.gov/datacatalog/data/USGS%3A69837e43b66b01367d7ec7c7"
)
MCS_SCIENCEBASE_URL = "https://www.sciencebase.gov/catalog/item/69837e43b66b01367d7ec7c7"
MCS_DOI_URL = "https://doi.org/10.5066/P1WKQ63T"
MYB_RELEASE_URL = "https://www.usgs.gov/media/files/rare-earths-2022-tables-only-release"
MYB_DOWNLOAD_URL = (
    "https://d9-wret.s3.us-west-2.amazonaws.com/assets/palladium/production/"
    "s3fs-public/media/files/myb1-2022-raree-ert.xlsx"
)

MCS_CHAPTERS = (
    "RARE EARTHS",
    "RARE EARTHS (Heavy)",
    "SCANDIUM",
    "YTTRIUM",
)
MCS_EXPECTED_CHAPTER_COUNTS = {
    "RARE EARTHS": 164,
    "RARE EARTHS (Heavy)": 51,
    "SCANDIUM": 30,
    "YTTRIUM": 41,
}

RAW_FILES = {
    "mcs_data": "usgs_mcs2026_commodities_data.csv",
    "mcs_metadata": "usgs_mcs2026_metadata.xml",
    "mcs_version_history": "usgs_mcs2026_version_history.txt",
    "mcs_rare_earths_pdf": "usgs_mcs2026_rare_earths.pdf",
    "mcs_heavy_pdf": "usgs_mcs2026_rare_earths_heavy.pdf",
    "mcs_scandium_pdf": "usgs_mcs2026_scandium.pdf",
    "mcs_yttrium_pdf": "usgs_mcs2026_yttrium.pdf",
    "myb_tables": "usgs_myb2022_rare_earths_tables.xlsx",
}
EXPECTED_RAW_SHA256 = {
    "mcs_data": "582a0aa231aea53d8a97dc8d1cd3dfa5f885cf3760353e3d029d7f0ae4fbaaf5",
    "mcs_metadata": "0588a74c5257484630cd41dbb60cbd9ef4862bd24b6660a762d1aecf3c73e15a",
    "mcs_version_history": "276d675d1d753697f611c385d0692c865f2c00b8a417c282ff2e8e1f84932e2e",
    "mcs_rare_earths_pdf": "0116a192336fec41c38d8e11dad553bb7703308c1fbd1f97dc14b75c7e7d9900",
    "mcs_heavy_pdf": "c9bf719946498b8ab90aceb901f94e63f2f5858c0efcba5b0232f3105eea5cc1",
    "mcs_scandium_pdf": "b68e156704a9f38a5f52cd30890901c3cdfe585bbcb7720b5e24238f0824d558",
    "mcs_yttrium_pdf": "047b71f4cb3faa57a2c203a98f3b1cf15530b7864e8ff9b61e5f5fa5ff356996",
    "myb_tables": "9f04f3418ab259e9565c154bb4833cc356203f79a20fa320b8e28041a0c4ca8e",
}
RAW_FILE_URLS = {
    "mcs_data": (
        "https://www.sciencebase.gov/catalog/file/get/69837e43b66b01367d7ec7c7?"
        "f=__disk__d3%2Fac%2F84%2Fd3ac8466552946c5e8caa2c2c6338d9e1aff655d"
    ),
    "mcs_metadata": (
        "https://www.sciencebase.gov/catalog/file/get/69837e43b66b01367d7ec7c7?"
        "f=__disk__59%2F9f%2F78%2F599f78dcdae6cfc820ec8410e8a0e684c28be9d2"
    ),
    "mcs_version_history": "https://pubs.usgs.gov/periodicals/mcs2026/versionHist.txt",
    "mcs_rare_earths_pdf": "https://pubs.usgs.gov/periodicals/mcs2026/mcs2026-rare-earths.pdf",
    "mcs_heavy_pdf": "https://pubs.usgs.gov/periodicals/mcs2026/mcs2026-rare-earths-heavy.pdf",
    "mcs_scandium_pdf": "https://pubs.usgs.gov/periodicals/mcs2026/mcs2026-scandium.pdf",
    "mcs_yttrium_pdf": "https://pubs.usgs.gov/periodicals/mcs2026/mcs2026-yttrium.pdf",
    "myb_tables": MYB_DOWNLOAD_URL,
}

MCS_COLUMNS = [
    "source_id",
    "source_row_number",
    "source_file",
    "mcs_chapter",
    "section",
    "commodity",
    "country",
    "statistics",
    "statistics_detail",
    "unit",
    "year",
    "raw_year",
    "raw_value",
    "raw_notes",
    "raw_is_critical_mineral_2025",
    "raw_other_notes",
    "current_value",
    "current_notes",
    "value",
    "value_low",
    "value_high",
    "comparator",
    "availability_status",
    "indicator_code",
    "is_estimated",
    "revision_action",
    "revision_version",
    "revision_source_file",
    "revision_page",
    "revision_note",
]

MCS_REVISION_COLUMNS = [
    "mcs_chapter",
    "section",
    "country",
    "statistics",
    "year",
    "raw_value",
    "current_value",
    "raw_notes",
    "current_notes",
    "revision_action",
    "revision_version",
    "revision_source_file",
    "revision_page",
    "revision_note",
]

MYB_COLUMNS = [
    "source_id",
    "source_file",
    "source_sheet",
    "source_row_number",
    "source_cell",
    "source_marker_cell",
    "source_country_label",
    "geography",
    "geography_code",
    "metric",
    "unit",
    "year",
    "raw_value",
    "display_value",
    "value",
    "availability_status",
    "raw_marker",
    "footnote_ids",
    "is_estimated",
    "is_revised",
    "data_status",
]

DATA_DICTIONARY_COLUMNS = ["dataset", "column", "definition"]

CRITICAL_RELIANCE_COLUMNS = [
    "source_id",
    "source_row_number",
    "source_file",
    "v3_mineral",
    "label",
    "scope",
    "mcs_chapter",
    "commodity",
    "statistics_detail",
    "year",
    "raw_value",
    "display_value",
    "value_pct",
    "value_low_pct",
    "comparator",
    "availability_status",
    "indicator_code",
    "is_estimated",
    "source_notes",
    "mapping_note",
]

OUTPUT_FILENAMES = {
    "mcs_observations": "usgs_mcs2026_observations.csv",
    "mcs_revision_audit": "usgs_mcs2026_revision_audit.csv",
    "myb_world_production": "usgs_myb2022_world_mine_production.csv",
    "data_dictionary": "usgs_publications_data_dictionary.csv",
    "metadata": "usgs_mcs2026_metadata.json",
    "critical_reliance": "usgs_mcs2026_critical_mineral_reliance.csv",
}

# A deliberately bounded map from MCS chapter measures to V3 analytical
# families. Production is not mapped because MCS units, stages, and content
# bases differ too much across commodities for one cross-mineral chart.
_CRITICAL_RELIANCE_MAP: tuple[dict[str, Any], ...] = (
    {"source_row_number": 2957, "v3_mineral": "natural_graphite", "label": "Natural graphite", "scope": "Natural graphite", "chapter": "GRAPHITE (NATURAL)", "commodity": "Graphite (Natural)", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "100"},
    {"source_row_number": 740, "v3_mineral": "aluminum", "label": "Bauxite", "scope": "Bauxite, not alumina or aluminum metal", "chapter": "BAUXITE AND ALUMINA", "commodity": "Bauxite", "detail": "Net import reliance as a percentage of apparent consumption", "expected": ">75"},
    {"source_row_number": 395, "v3_mineral": "antimony", "label": "Antimony", "scope": "Oxide and unwrought metal or powder", "chapter": "ANTIMONY", "commodity": "Antimony", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "91"},
    {"source_row_number": 993, "v3_mineral": "bismuth", "label": "Bismuth", "scope": "Bismuth", "chapter": "BISMUTH", "commodity": "Bismuth", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "92"},
    {"source_row_number": 1484, "v3_mineral": "chromium", "label": "Chromium", "scope": "Chromium content", "chapter": "CHROMIUM", "commodity": "Chromium", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "79"},
    {"source_row_number": 1844, "v3_mineral": "cobalt", "label": "Cobalt", "scope": "Refined cobalt", "chapter": "COBALT", "commodity": "Cobalt", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "79"},
    {"source_row_number": 1971, "v3_mineral": "copper", "label": "Copper", "scope": "Refined copper", "chapter": "COPPER", "commodity": "Copper", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "57"},
    {"source_row_number": 4578, "v3_mineral": "manganese", "label": "Manganese", "scope": "Manganese content", "chapter": "MANGANESE", "commodity": "Manganese", "detail": "Net import reliance as a percentage of apparent consumption, manganese content", "expected": "100"},
    {"source_row_number": 5040, "v3_mineral": "nickel", "label": "Nickel", "scope": "Total consumption including scrap", "chapter": "NICKEL", "commodity": "Nickel", "detail": "Net import reliance as a percentage of total apparent consumption", "expected": "41"},
    {"source_row_number": 7485, "v3_mineral": "tantalum", "label": "Tantalum", "scope": "Tantalum content", "chapter": "TANTALUM", "commodity": "Tantalum", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "100"},
    {"source_row_number": 7800, "v3_mineral": "tin", "label": "Tin", "scope": "Refined tin", "chapter": "TIN", "commodity": "Tin", "detail": "Net import reliance as a percentage of apparent consumption, refined tin", "expected": "77"},
    {"source_row_number": 8200, "v3_mineral": "tungsten", "label": "Tungsten", "scope": "Tungsten content", "chapter": "TUNGSTEN", "commodity": "Tungsten", "detail": "Net import reliance as a percentage of apparent consumption", "expected": ">50"},
    {"source_row_number": 8676, "v3_mineral": "zinc", "label": "Zinc", "scope": "Refined zinc", "chapter": "ZINC", "commodity": "Zinc", "detail": "Net import reliance as a percentage of apparent consumption: Refined zinc", "expected": "73"},
    {"source_row_number": 6064, "v3_mineral": "rare_earths", "label": "Rare earths", "scope": "Compounds and metals", "chapter": "RARE EARTHS", "commodity": "Rare Earths", "detail": "Net import reliance as a percentage of apparent consumption: Compounds and metals", "expected": "67"},
    {"source_row_number": 6164, "v3_mineral": "heavy_rare_earths", "label": "Heavy rare earths", "scope": "Compounds and metals; excludes yttrium", "chapter": "RARE EARTHS (Heavy)", "commodity": "Rare Earths (Heavy)", "detail": "Net import reliance as a percentage of apparent consumption, compounds and metals", "expected": "100"},
    {"source_row_number": 6546, "v3_mineral": "scandium", "label": "Scandium", "scope": "Scandium", "chapter": "SCANDIUM", "commodity": "Scandium", "detail": "Net import reliance as a percentage of apparent consumption", "expected": "100"},
    {"source_row_number": 8527, "v3_mineral": "yttrium", "label": "Yttrium", "scope": "Compounds and metals", "chapter": "YTTRIUM", "commodity": "Yttrium", "detail": "Net import relianceas a percentage of apparent consumption", "expected": "100"},
)

_MCS_RAW_HEADERS = [
    "MCS chapter",
    "Section",
    "Commodity",
    "Country",
    "Statistics",
    "Statistics_detail",
    "Unit",
    "Year",
    "Value",
    "Notes",
    "Is critical mineral 2025",
    "Other notes",
]

_WORLD_SECTION = "World Mine Production and Reserves"
_VERSION_HISTORY_SOURCE = f"data/raw/{RAW_FILES['mcs_version_history']}"
_CURRENT_PDF_SOURCE = f"data/raw/{RAW_FILES['mcs_rare_earths_pdf']}"
_INDIA_CURRENT_NOTE = (
    "See Appendix C for resource and reserve definitions and information concerning data sources."
    " -- A 2015 report from OSCOM indicated that monazite reserves from their operations were "
    "256,000 tons; rare-earth reserves were not reported."
)

# Keys are (chapter, section, country, statistic, year).  Expected raw content
# makes each correction fail loudly if USGS replaces the frozen ScienceBase CSV.
_CURRENT_REVISIONS: dict[tuple[str, str, str, str, str], dict[str, str]] = {
    ("RARE EARTHS", _WORLD_SECTION, "Brazil", "Reserves", "2025"): {
        "expected_raw_value": "21,000,000",
        "current_value": "11,000,000",
        "action": "replace_value",
        "version": "1.3",
        "note": (
            "MCS 2026 version 1.3 revised Brazil reserves from 21,000,000 to "
            "11,000,000 metric tons."
        ),
    },
    ("RARE EARTHS", _WORLD_SECTION, "World total", "Reserves", "2025"): {
        "expected_raw_value": ">85,000,000",
        "current_value": ">75,000,000",
        "action": "replace_value",
        "version": "1.3",
        "note": (
            "MCS 2026 version 1.3 revised the rounded world reserve lower bound "
            "from greater than 85,000,000 to greater than 75,000,000 metric tons."
        ),
    },
    ("RARE EARTHS", _WORLD_SECTION, "China", "Production", "2024"): {
        "expected_raw_value": "270,000",
        "expected_raw_notes": "Estimated. -- Production quota; does not include undocumented production.",
        "current_value": "270,000",
        "current_notes": "Estimated.",
        "action": "remove_superseded_note",
        "version": "1.1",
        "note": (
            "MCS 2026 version 1.1 removed footnote 14 from China's 2024 mine production; "
            "the production value did not change."
        ),
    },
    ("RARE EARTHS", _WORLD_SECTION, "India", "Reserves", "2025"): {
        "expected_raw_value": "NA",
        "current_value": "NA",
        "current_notes": _INDIA_CURRENT_NOTE,
        "action": "add_reassigned_note",
        "version": "1.1",
        "note": (
            "MCS 2026 version 1.1 assigned revised footnote 14 to India's reserves. "
            "The reserve value remains not available."
        ),
    },
}

_MYB_COUNTRIES: dict[str, tuple[str, str, bool, tuple[str, ...]]] = {
    "Australiae": ("Australia", "AUS", True, ()),
    "Brazile": ("Brazil", "BRA", True, ()),
    "Burmae": ("Burma", "MMR", True, ()),
    "Burundi": ("Burundi", "BDI", False, ()),
    "China3": ("China", "CHN", False, ("3",)),
    "Indiae, 4": ("India", "IND", True, ("4",)),
    "Madagascare": ("Madagascar", "MDG", True, ()),
    "Malaysiae": ("Malaysia", "MYS", True, ()),
    "Russia": ("Russia", "RUS", False, ()),
    "Thailande, 5": ("Thailand", "THA", True, ("5",)),
    "United States": ("United States", "USA", False, ()),
    "Vietname, 5": ("Vietnam", "VNM", True, ("5",)),
    "Total": ("World total", "WLD", False, ()),
}

_MYB_FOOTNOTES = {
    "1": (
        "Table includes data available through April 1, 2024. All data are reported unless otherwise "
        "noted; totals may include estimated data. Totals, U.S. data, and estimated data are rounded "
        "to no more than three significant digits; may not add to totals shown."
    ),
    "2": (
        "Other possible producers are omitted where available information was inadequate for reliable estimates."
    ),
    "3": "Official production quota. Illegal production could not be quantified.",
    "4": "India's Department of Atomic Energy did not disclose monazite production data.",
    "5": "Rare-earth oxide content of exports.",
    "6": "Includes bastnaesite only. Excludes monazite from heavy mineral sands.",
}


def _sha256(path: Path) -> str:
    """Return a lowercase hexadecimal SHA-256 digest for *path*."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _as_number(text: str) -> int | float:
    """Convert a comma-formatted decimal string to an int when lossless."""

    number = float(text.replace(",", "").strip())
    return int(number) if number.is_integer() else number


def parse_published_value(raw_value: Any) -> dict[str, Any]:
    """Parse one USGS display value without inventing a midpoint.

    Ranges use ``value_low`` and ``value_high``; lower-bound observations use
    only ``value_low``.  ``E`` is the USGS net-exporter indicator, not an
    estimate marker.  An em dash or double hyphen is an explicit zero.
    """

    text = "" if raw_value is None else str(raw_value).strip()
    blank = {
        "value": "",
        "value_low": "",
        "value_high": "",
        "comparator": "",
        "availability_status": "not_available",
        "indicator_code": "",
    }
    if not text or text.upper() == "NA":
        return blank
    if text in {"—", "--"}:
        return {
            "value": 0,
            "value_low": "",
            "value_high": "",
            "comparator": "exact",
            "availability_status": "explicit_zero",
            "indicator_code": "",
        }
    if text.upper() == "E":
        return {
            **blank,
            "availability_status": "indicator",
            "indicator_code": "net_exporter",
        }
    if text.startswith(">"):
        return {
            "value": "",
            "value_low": _as_number(text[1:]),
            "value_high": "",
            "comparator": "greater_than",
            "availability_status": "available",
            "indicator_code": "",
        }
    range_match = re.fullmatch(
        r"\s*([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:–|-)\s*([0-9][0-9,]*(?:\.[0-9]+)?)\s*",
        text,
    )
    if range_match:
        return {
            "value": "",
            "value_low": _as_number(range_match.group(1)),
            "value_high": _as_number(range_match.group(2)),
            "comparator": "range",
            "availability_status": "available",
            "indicator_code": "",
        }
    try:
        number = _as_number(text)
    except ValueError as exc:
        raise ValueError(f"unrecognized USGS value {text!r}") from exc
    return {
        "value": number,
        "value_low": "",
        "value_high": "",
        "comparator": "exact",
        "availability_status": "available",
        "indicator_code": "",
    }


def parse_mcs_2026(raw_dir: Path = RAW_DIR) -> list[dict[str, Any]]:
    """Read and normalize the four rare-earth MCS chapters (286 rows)."""

    path = Path(raw_dir) / RAW_FILES["mcs_data"]
    rows: list[dict[str, Any]] = []
    applied_revisions: Counter[tuple[str, str, str, str, str]] = Counter()
    with path.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != _MCS_RAW_HEADERS:
            raise ValueError(f"{path.name}: unexpected headers {reader.fieldnames!r}")
        for source_row_number, raw in enumerate(reader, start=2):
            if raw["MCS chapter"] not in MCS_CHAPTERS:
                continue
            key = (
                raw["MCS chapter"],
                raw["Section"],
                raw["Country"],
                raw["Statistics"],
                raw["Year"],
            )
            revision = _CURRENT_REVISIONS.get(key)
            current_value = raw["Value"]
            current_notes = raw["Notes"]
            revision_fields = {
                "revision_action": "none",
                "revision_version": "",
                "revision_source_file": "",
                "revision_page": "",
                "revision_note": "",
            }
            if revision:
                if raw["Value"] != revision["expected_raw_value"]:
                    raise ValueError(
                        f"{path.name}:{source_row_number}: revision raw value changed: "
                        f"{raw['Value']!r} != {revision['expected_raw_value']!r}"
                    )
                expected_notes = revision.get("expected_raw_notes")
                if expected_notes is not None and raw["Notes"] != expected_notes:
                    raise ValueError(
                        f"{path.name}:{source_row_number}: revision raw note changed: {raw['Notes']!r}"
                    )
                current_value = revision.get("current_value", current_value)
                current_notes = revision.get("current_notes", current_notes)
                revision_fields = {
                    "revision_action": revision["action"],
                    "revision_version": revision["version"],
                    "revision_source_file": _VERSION_HISTORY_SOURCE,
                    "revision_page": "153",
                    "revision_note": revision["note"],
                }
                applied_revisions[key] += 1

            parsed = parse_published_value(current_value)
            raw_year = raw["Year"]
            year: int | str = int(raw_year) if raw_year.isdigit() else raw_year
            rows.append(
                {
                    "source_id": "usgs_mcs2026_v1_3",
                    "source_row_number": source_row_number,
                    "source_file": f"data/raw/{path.name}",
                    "mcs_chapter": raw["MCS chapter"],
                    "section": raw["Section"],
                    "commodity": raw["Commodity"],
                    "country": raw["Country"],
                    "statistics": raw["Statistics"],
                    "statistics_detail": raw["Statistics_detail"],
                    "unit": raw["Unit"],
                    "year": year,
                    "raw_year": raw_year,
                    "raw_value": raw["Value"],
                    "raw_notes": raw["Notes"],
                    "raw_is_critical_mineral_2025": raw["Is critical mineral 2025"],
                    "raw_other_notes": raw["Other notes"],
                    "current_value": current_value,
                    "current_notes": current_notes,
                    **parsed,
                    "is_estimated": "Estimated." in current_notes,
                    **revision_fields,
                }
            )

    chapter_counts = Counter(row["mcs_chapter"] for row in rows)
    if len(rows) != 286 or dict(chapter_counts) != MCS_EXPECTED_CHAPTER_COUNTS:
        raise ValueError(
            f"{path.name}: expected 286 selected rows {MCS_EXPECTED_CHAPTER_COUNTS}; "
            f"found {len(rows)} {dict(chapter_counts)}"
        )
    if set(applied_revisions) != set(_CURRENT_REVISIONS) or any(
        count != 1 for count in applied_revisions.values()
    ):
        raise ValueError(f"{path.name}: current-PDF revisions did not match exactly once")
    return rows


def parse_mcs_2026_critical_reliance(raw_dir: Path = RAW_DIR) -> list[dict[str, Any]]:
    """Extract 17 fail-loud, scope-labeled 2025 critical-mineral reliance rows."""

    path = Path(raw_dir) / RAW_FILES["mcs_data"]
    expected_by_row = {item["source_row_number"]: item for item in _CRITICAL_RELIANCE_MAP}
    observed: dict[int, dict[str, Any]] = {}
    with path.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != _MCS_RAW_HEADERS:
            raise ValueError(f"{path.name}: unexpected headers {reader.fieldnames!r}")
        for source_row_number, raw in enumerate(reader, start=2):
            mapping = expected_by_row.get(source_row_number)
            if mapping is None:
                continue
            expected_fields = {
                "MCS chapter": mapping["chapter"],
                "Section": "Salient Statistics—United States",
                "Commodity": mapping["commodity"],
                "Country": "United States",
                "Statistics": "Net import reliance",
                "Statistics_detail": mapping["detail"],
                "Unit": "percent",
                "Year": "2025",
                "Value": mapping["expected"],
                "Is critical mineral 2025": "Yes",
            }
            changed = {
                field: (raw[field], expected)
                for field, expected in expected_fields.items()
                if raw[field] != expected
            }
            if changed:
                raise ValueError(
                    f"{path.name}:{source_row_number}: mapped reliance row changed: {changed}"
                )
            parsed = parse_published_value(raw["Value"])
            if parsed["availability_status"] != "available" or parsed["indicator_code"]:
                raise ValueError(
                    f"{path.name}:{source_row_number}: expected numeric or lower-bound reliance"
                )
            observed[source_row_number] = {
                "source_id": "usgs_mcs2026_critical_mineral_reliance_2025",
                "source_row_number": source_row_number,
                "source_file": f"data/raw/{path.name}",
                "v3_mineral": mapping["v3_mineral"],
                "label": mapping["label"],
                "scope": mapping["scope"],
                "mcs_chapter": raw["MCS chapter"],
                "commodity": raw["Commodity"],
                "statistics_detail": raw["Statistics_detail"],
                "year": 2025,
                "raw_value": raw["Value"],
                "display_value": f"{raw['Value']}%",
                "value_pct": parsed["value"],
                "value_low_pct": parsed["value_low"],
                "comparator": parsed["comparator"],
                "availability_status": parsed["availability_status"],
                "indicator_code": parsed["indicator_code"],
                "is_estimated": "Estimated." in raw["Notes"],
                "source_notes": raw["Notes"],
                "mapping_note": (
                    "MCS chapter-level dependency measure mapped to a V3 analytical family; "
                    "not derived from HTS trade rows and not a China share."
                ),
            }
    if set(observed) != set(expected_by_row):
        missing = sorted(set(expected_by_row) - set(observed))
        raise ValueError(f"{path.name}: missing mapped critical-mineral reliance rows {missing}")
    rows = [observed[item["source_row_number"]] for item in _CRITICAL_RELIANCE_MAP]
    if len(rows) != 17 or not all(row["is_estimated"] for row in rows):
        raise ValueError(f"{path.name}: expected 17 estimated critical-mineral reliance rows")
    return rows


def _display_myb_value(raw_value: Any) -> str:
    """Render a T8 value as displayed in the source table."""

    if isinstance(raw_value, bool):
        return str(raw_value)
    if isinstance(raw_value, int):
        return f"{raw_value:,}"
    if isinstance(raw_value, float) and raw_value.is_integer():
        return f"{int(raw_value):,}"
    return "" if raw_value is None else str(raw_value)


def parse_myb_2022_t8(raw_dir: Path = RAW_DIR) -> list[dict[str, Any]]:
    """Normalize only MYB 2022 table T8 into 65 country-year rows."""

    path = Path(raw_dir) / RAW_FILES["myb_tables"]
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name.*")
        formula_workbook = load_workbook(path, data_only=False, read_only=False)
        value_workbook = load_workbook(path, data_only=True, read_only=False)
    if formula_workbook.sheetnames != ["Note", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8"]:
        raise ValueError(f"{path.name}: unexpected sheet list {formula_workbook.sheetnames!r}")
    formula_sheet = formula_workbook["T8"]
    sheet = value_workbook["T8"]
    if formula_sheet["A1"].value != "TABLE 8" or not str(formula_sheet["A2"].value).startswith(
        "RARE EARTHS: WORLD MINE PRODUCTION"
    ):
        raise ValueError(f"{path.name}: T8 title changed")
    year_columns = {3: 2018, 5: 2019, 7: 2020, 9: 2021, 11: 2022}
    if {column: sheet.cell(6, column).value for column in year_columns} != year_columns:
        raise ValueError(f"{path.name}: T8 year headings changed")
    for row_number in range(7, 20):
        for column in year_columns:
            if formula_sheet.cell(row_number, column).data_type == "f":
                raise ValueError(f"{path.name}: unexpected formula in {formula_sheet.cell(row_number, column).coordinate}")

    rows: list[dict[str, Any]] = []
    for row_number in range(7, 20):
        source_label = str(sheet.cell(row_number, 1).value)
        if source_label not in _MYB_COUNTRIES:
            raise ValueError(f"{path.name}: unrecognized T8 geography label {source_label!r}")
        geography, geography_code, row_estimated, row_footnotes = _MYB_COUNTRIES[source_label]
        for value_column, year in year_columns.items():
            value_cell = sheet.cell(row_number, value_column)
            marker_cell = sheet.cell(row_number, value_column + 1)
            raw_value = value_cell.value
            raw_marker = "" if marker_cell.value is None else str(marker_cell.value)
            if raw_marker not in {"", "e", "r", "6"}:
                raise ValueError(
                    f"{path.name}: unexpected T8 marker {raw_marker!r} at {marker_cell.coordinate}"
                )
            parsed = parse_published_value(raw_value)
            if parsed["value_low"] != "" or parsed["value_high"] != "" or parsed["indicator_code"]:
                raise ValueError(f"{path.name}: unsupported T8 value at {value_cell.coordinate}")
            is_estimated = row_estimated or raw_marker == "e"
            is_revised = raw_marker == "r"
            footnote_ids = [*row_footnotes]
            if raw_marker == "6":
                footnote_ids.append("6")
            if parsed["availability_status"] == "explicit_zero":
                data_status = "explicit_zero"
            elif is_estimated and is_revised:
                data_status = "estimated_revised"
            elif is_revised:
                data_status = "revised"
            elif is_estimated:
                data_status = "estimated"
            else:
                data_status = "reported"
            rows.append(
                {
                    "source_id": "usgs_myb2022_t8",
                    "source_file": f"data/raw/{path.name}",
                    "source_sheet": "T8",
                    "source_row_number": row_number,
                    "source_cell": value_cell.coordinate,
                    "source_marker_cell": marker_cell.coordinate,
                    "source_country_label": source_label,
                    "geography": geography,
                    "geography_code": geography_code,
                    "metric": "mine_production",
                    "unit": "metric_tons_reo_equivalent",
                    "year": year,
                    "raw_value": raw_value,
                    "display_value": _display_myb_value(raw_value),
                    "value": parsed["value"],
                    "availability_status": parsed["availability_status"],
                    "raw_marker": raw_marker,
                    "footnote_ids": "|".join(footnote_ids),
                    "is_estimated": is_estimated,
                    "is_revised": is_revised,
                    "data_status": data_status,
                }
            )
    if len(rows) != 65 or Counter(row["year"] for row in rows) != Counter(
        {2018: 13, 2019: 13, 2020: 13, 2021: 13, 2022: 13}
    ):
        raise ValueError(f"{path.name}: expected 65 T8 observations")
    return rows


def verify_frozen_inputs(raw_dir: Path = RAW_DIR) -> list[dict[str, Any]]:
    """Verify all frozen inputs and return file-level provenance records."""

    records: list[dict[str, Any]] = []
    for source_key, filename in RAW_FILES.items():
        path = Path(raw_dir) / filename
        if not path.is_file():
            raise FileNotFoundError(path)
        actual_hash = _sha256(path)
        expected_hash = EXPECTED_RAW_SHA256[source_key]
        if actual_hash != expected_hash:
            raise ValueError(f"{filename}: SHA-256 {actual_hash} != frozen {expected_hash}")
        records.append(
            {
                "source_key": source_key,
                "file": f"data/raw/{filename}",
                "bytes": path.stat().st_size,
                "sha256": actual_hash,
                "url": RAW_FILE_URLS[source_key],
            }
        )
    return records


def build_usgs_publications_metadata(
    mcs_rows: list[dict[str, Any]],
    myb_rows: list[dict[str, Any]],
    critical_reliance_rows: list[dict[str, Any]],
    raw_dir: Path = RAW_DIR,
) -> dict[str, Any]:
    """Build machine-readable provenance for both normalized USGS datasets."""

    raw_files = verify_frozen_inputs(raw_dir)
    return {
        "schema_version": "1.1.0",
        "source_id": "usgs_rare_earths_publications",
        "title": "USGS rare-earth statistics and information hub ingest",
        "retrieved_at": RETRIEVED_AT,
        "source_hub_url": SOURCE_HUB_URL,
        "raw_files": raw_files,
        "datasets": {
            "mcs2026": {
                "source_id": "usgs_mcs2026_v1_3",
                "title": "Mineral Commodity Summaries 2026 rare-earth chapters",
                "publication_date": "2026-02-06",
                "current_version": "1.3",
                "current_version_date": "2026-05-27",
                "data_release_url": MCS_DATA_RELEASE_URL,
                "sciencebase_url": MCS_SCIENCEBASE_URL,
                "doi_url": MCS_DOI_URL,
                "source_encoding": "cp1252",
                "chapters": list(MCS_CHAPTERS),
                "chapter_row_counts": dict(Counter(row["mcs_chapter"] for row in mcs_rows)),
                "row_count": len(mcs_rows),
                "revision_count": sum(row["revision_action"] != "none" for row in mcs_rows),
                "revision_policy": (
                    "Raw ScienceBase fields are retained. Four audited version-history changes are "
                    "applied only to separate current-view and parsed fields."
                ),
                "current_pdf_file": _CURRENT_PDF_SOURCE,
                "revision_history_file": _VERSION_HISTORY_SOURCE,
            },
            "myb2022_t8": {
                "source_id": "usgs_myb2022_t8",
                "title": "Rare Earths 2022 tables-only release, table T8",
                "release_page_url": MYB_RELEASE_URL,
                "download_url": MYB_DOWNLOAD_URL,
                "source_sheet": "T8",
                "coverage": "2018–2022",
                "data_available_through": "2024-04-01",
                "row_count": len(myb_rows),
                "table_scope": "World mine production by country or locality, REO equivalent.",
                "excluded_tables": (
                    "T1–T7 remain in the frozen workbook and are not normalized here, preventing "
                    "MYB partner-trade tables from being conflated with the DataWeb trade record."
                ),
                "footnotes": _MYB_FOOTNOTES,
            },
            "critical_mineral_reliance_2025": {
                "source_id": "usgs_mcs2026_critical_mineral_reliance_2025",
                "title": "Selected U.S. critical-mineral net-import-reliance indicators",
                "coverage": [2025, 2025],
                "row_count": len(critical_reliance_rows),
                "selection_rule": (
                    "Seventeen exact MCS chapter measures mapped to V3 mineral families only where "
                    "the source flags the row as a 2025 critical mineral and the net-import-reliance "
                    "percentage and material scope can be labeled directly."
                ),
                "comparability_warning": (
                    "Scopes and apparent-consumption denominators differ; values must not be averaged, "
                    "summed, or interpreted as China shares."
                ),
            },
        },
        "outputs": {
            "mcs_observations": {
                "file": f"data/processed/{OUTPUT_FILENAMES['mcs_observations']}",
                "rows": len(mcs_rows),
            },
            "mcs_revision_audit": {
                "file": f"data/processed/{OUTPUT_FILENAMES['mcs_revision_audit']}",
                "rows": sum(row["revision_action"] != "none" for row in mcs_rows),
            },
            "myb_world_production": {
                "file": f"data/processed/{OUTPUT_FILENAMES['myb_world_production']}",
                "rows": len(myb_rows),
            },
            "data_dictionary": {
                "file": f"data/processed/{OUTPUT_FILENAMES['data_dictionary']}"
            },
            "critical_reliance": {
                "file": f"data/processed/{OUTPUT_FILENAMES['critical_reliance']}",
                "rows": len(critical_reliance_rows),
            },
        },
        "caveats": [
            "MCS 2026 is a 2026 publication whose tabular observations chiefly cover 2021–2025.",
            "Mine production is upstream supply context, not a measure of national access or control.",
            "MCS import-source percentages describe direct or shipping sources, not necessarily mine origin.",
            "Rare-earth imports from Estonia, Japan, and Malaysia may derive from concentrates or intermediates produced in Australia, China, or elsewhere.",
            "Heavy-rare-earth, scandium, and yttrium source shares exclude material contained in value-added intermediates and finished goods.",
            "The MCS source does not publish dysprosium or thulium import-source shares; absence is not zero.",
            "The MCS ScienceBase CSV reflects the initial release; the current view explicitly applies versions 1.1 and 1.3 without changing the raw fields.",
            "MYB 2022 T8 and MCS 2026 do not provide a comparable 2023 world mine-production table in this ingest; the site must display that year as missing.",
        ],
    }


def build_usgs_publications_context(
    mcs_rows: list[dict[str, Any]],
    myb_rows: list[dict[str, Any]],
    critical_reliance_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Return the compact, claim-safe subset intended for the static site."""

    production: dict[tuple[int, str], int | float] = {}
    for row in myb_rows:
        if row["availability_status"] in {"available", "explicit_zero"}:
            production[(int(row["year"]), row["geography_code"])] = row["value"]
    for row in mcs_rows:
        if (
            row["mcs_chapter"] == "RARE EARTHS"
            and row["section"] == _WORLD_SECTION
            and row["statistics"] == "Production"
            and row["year"] in {2024, 2025}
            and row["country"] in {"China", "United States", "World total"}
            and row["availability_status"] == "available"
        ):
            code = {"China": "CHN", "United States": "USA", "World total": "WLD"}[
                row["country"]
            ]
            production[(int(row["year"]), code)] = row["value"]

    share_series: list[dict[str, Any]] = []
    for year in range(2018, 2026):
        if year == 2023:
            share_series.append(
                {
                    "year": year,
                    "china_production": None,
                    "us_production": None,
                    "world_production": None,
                    "china_share_percent": None,
                    "us_share_percent": None,
                    "data_status": "missing",
                    "missing_data_label": "No comparable table ingested for 2023",
                    "source_id": None,
                }
            )
            continue
        china = production[(year, "CHN")]
        united_states = production[(year, "USA")]
        world = production[(year, "WLD")]
        share_series.append(
            {
                "year": year,
                "china_production": china,
                "us_production": united_states,
                "world_production": world,
                "china_share_percent": round(china / world * 100, 1),
                "us_share_percent": round(united_states / world * 100, 1),
                "data_status": "available",
                "missing_data_label": None,
                "source_id": "usgs_myb2022_t8" if year <= 2022 else "usgs_mcs2026_v1_3",
            }
        )
    latest = share_series[-1]

    import_sources = [
        {
            "chapter": row["mcs_chapter"],
            "commodity": row["commodity"],
            "material": row["statistics_detail"],
            "country": row["country"],
            "period": row["raw_year"],
            "share_percent": row["value"],
            "notes": row["current_notes"],
        }
        for row in mcs_rows
        if row["section"] == "Import Sources"
    ]
    if len(import_sources) != 23:
        raise ValueError(f"MCS import-source snapshot expected 23 rows; found {len(import_sources)}")

    def source_shares(chapter: str) -> dict[str, dict[str, int | float]]:
        """Group direct-source shares first by material and then by country."""

        grouped: dict[str, dict[str, int | float]] = {}
        for observation in import_sources:
            if observation["chapter"] == chapter:
                grouped.setdefault(str(observation["material"]), {})[str(observation["country"])] = observation[
                    "share_percent"
                ]
        return grouped

    def observation(
        chapter: str,
        statistic: str,
        year: int,
        *,
        detail: str | None = None,
        country: str = "United States",
        section: str = "Salient Statistics—United States",
    ) -> dict[str, Any]:
        """Return one exact normalized MCS observation or fail loudly."""

        matches = [
            row
            for row in mcs_rows
            if row["mcs_chapter"] == chapter
            and row["section"] == section
            and row["country"] == country
            and row["statistics"] == statistic
            and row["year"] == year
            and (detail is None or row["statistics_detail"] == detail)
        ]
        if len(matches) != 1:
            raise ValueError(
                "MCS site observation expected exactly one row: "
                f"{chapter}/{section}/{country}/{statistic}/{detail}/{year}; found {len(matches)}"
            )
        return matches[0]

    def compact_observation(row: Mapping[str, Any]) -> dict[str, Any]:
        """Keep numeric state and the exact publication display token together."""

        return {
            "display": row["current_value"],
            "value": row["value"] if row["value"] != "" else None,
            "value_low": row["value_low"] if row["value_low"] != "" else None,
            "value_high": row["value_high"] if row["value_high"] != "" else None,
            "comparator": row["comparator"] or None,
            "availability_status": row["availability_status"],
            "indicator_code": row["indicator_code"] or None,
            "unit": row["unit"],
        }

    us_balance_series: list[dict[str, Any]] = []
    reliance_scopes = (
        (
            "rare_earths_compounds_metals",
            "Rare-earth compounds and metals",
            "RARE EARTHS",
            "Net import reliance as a percentage of apparent consumption: Compounds and metals",
        ),
        (
            "heavy_rare_earths_compounds_metals",
            "Heavy rare-earth compounds and metals",
            "RARE EARTHS (Heavy)",
            "Net import reliance as a percentage of apparent consumption, compounds and metals",
        ),
        (
            "scandium",
            "Scandium",
            "SCANDIUM",
            "Net import reliance as a percentage of apparent consumption",
        ),
        (
            "yttrium",
            "Yttrium",
            "YTTRIUM",
            "Net import relianceas a percentage of apparent consumption",
        ),
    )
    reliance_series = [
        {"scope_id": scope_id, "label": label, "values": []}
        for scope_id, label, _chapter, _detail in reliance_scopes
    ]
    stage_measurement_specs = (
        ("mineral_concentrate_production", "Mineral-concentrate production"),
        ("compounds_metals_production", "Compounds-and-metals production"),
        ("compound_imports", "Compound imports"),
        ("apparent_consumption_compounds_metals", "Apparent consumption, compounds and metals"),
        ("mine_mill_employment", "Mine-and-mill employment"),
    )
    stage_measurement_rows: dict[str, list[dict[str, Any]]] = {
        measurement_id: [] for measurement_id, _label in stage_measurement_specs
    }
    for year in range(2021, 2026):
        concentrate_production = observation(
            "RARE EARTHS", "Production", year, detail="Production: Mineral concentrates"
        )
        downstream_production = observation(
            "RARE EARTHS", "Production", year, detail="Production: Compounds and metals"
        )
        compound_imports = observation(
            "RARE EARTHS", "Import", year, detail="Imports: Compounds"
        )
        apparent_consumption = observation(
            "RARE EARTHS",
            "Consumption",
            year,
            detail="Consumption, apparent, compounds and metals",
        )
        employment = observation(
            "RARE EARTHS",
            "Employment",
            year,
            detail="Employment, mine and mill, annual average, number",
        )
        stage_measurement_rows["mineral_concentrate_production"].append(concentrate_production)
        stage_measurement_rows["compounds_metals_production"].append(downstream_production)
        stage_measurement_rows["compound_imports"].append(compound_imports)
        stage_measurement_rows["apparent_consumption_compounds_metals"].append(apparent_consumption)
        stage_measurement_rows["mine_mill_employment"].append(employment)
        concentrate_reliance = observation(
            "RARE EARTHS",
            "Net import reliance",
            year,
            detail="Net import reliance as a percentage of apparent consumption: Mineral concentrates",
        )
        downstream_reliance = observation(
            "RARE EARTHS",
            "Net import reliance",
            year,
            detail="Net import reliance as a percentage of apparent consumption: Compounds and metals",
        )
        us_balance_series.append(
            {
                "year": year,
                "mineral_concentrate_production": concentrate_production["value"],
                "compounds_metals_production": downstream_production["value"],
                "compound_imports": compound_imports["value"],
                "apparent_consumption_compounds_metals": apparent_consumption["value"],
                "mine_mill_employment": employment["value"],
                "compounds_metals_net_import_reliance": compact_observation(downstream_reliance),
                "mineral_concentrate_trade_status": compact_observation(concentrate_reliance),
            }
        )
        for target, (_scope_id, _label, chapter, detail) in zip(reliance_series, reliance_scopes):
            target["values"].append(
                {
                    "year": year,
                    **compact_observation(
                        observation(chapter, "Net import reliance", year, detail=detail)
                    ),
                }
            )

    measurement_provenance = []
    for measurement_id, label in stage_measurement_specs:
        measurement_rows = stage_measurement_rows[measurement_id]
        note_groups: list[dict[str, Any]] = []
        for row in measurement_rows:
            note = str(row["current_notes"] or "").strip()
            if not note:
                continue
            matching_group = next(
                (group for group in note_groups if group["text"] == note),
                None,
            )
            if matching_group is None:
                note_groups.append({"years": [row["year"]], "text": note})
            else:
                matching_group["years"].append(row["year"])
        units = {row["unit"] for row in measurement_rows}
        if len(units) != 1:
            raise ValueError(f"MCS stage measurement changed unit: {measurement_id}/{sorted(units)}")
        measurement_provenance.append(
            {
                "measurement_id": measurement_id,
                "label": label,
                "unit": measurement_rows[0]["unit"],
                "source_rows": [
                    {"year": row["year"], "source_row_number": row["source_row_number"]}
                    for row in measurement_rows
                ],
                "estimated_years": [
                    row["year"] for row in measurement_rows if row["is_estimated"]
                ],
                "source_notes": note_groups,
            }
        )

    reserves_2025 = []
    for country, geography_code in (
        ("United States", "USA"),
        ("China", "CHN"),
        ("World total", "WLD"),
    ):
        row = observation(
            "RARE EARTHS",
            "Reserves",
            2025,
            detail="Reserves" if country != "World total" else "Reserves: rounded",
            country=country,
            section=_WORLD_SECTION,
        )
        reserves_2025.append(
            {
                "geography": country,
                "geography_code": geography_code,
                "unit_basis": "metric_tons_reo_equivalent",
                **compact_observation(row),
            }
        )

    latest_us_balance = us_balance_series[-1]
    us_statistical_baseline = {
        "status": "loaded",
        "coverage": [2021, 2025],
        "mass_unit": "metric_tons_reo_equivalent",
        "series": us_balance_series,
        "latest": latest_us_balance,
        "net_import_reliance": reliance_series,
        "measurement_provenance": measurement_provenance,
        "reserves_2025": reserves_2025,
        "source_vintage": "USGS Mineral Commodity Summaries 2026, version 1.3",
        "source_observation_year": 2025,
        "warnings": [
            "MCS 2026 is a publication vintage; the newest observations are estimates for 2025.",
            "Mineral concentrates and compounds or metals are different processing stages and must not be summed.",
            "Compound imports are a narrower product scope than apparent consumption of compounds and metals.",
            "The E indicator means net exporter; it is not an estimated percentage.",
            "Reserve figures describe geologic availability under USGS definitions, not ownership or assured access.",
        ],
    }
    indicator_notes = {
        "nickel": (
            "Includes stainless-steel and alloy scrap; USGS states that reliance would be nearly "
            "100% if scrap were excluded."
        ),
        "heavy_rare_earths": "Quantitative export data were unavailable.",
        "scandium": "Quantitative export data were unavailable.",
        "rare_earths": "Includes yttrium but excludes most scandium.",
        "yttrium": "General rare-earth trade data already include yttrium.",
    }
    us_official_baseline = {
        "status": "loaded",
        "observation_year": 2025,
        "publication_year": 2026,
        "publication_version": "MCS 2026 v1.3",
        "unit": "percent_of_apparent_consumption",
        "indicators": [
            {
                "id": row["v3_mineral"],
                "label": row["label"],
                "mineral_id": row["v3_mineral"],
                "scope": row["scope"],
                "value_pct": row["value_pct"] if row["value_pct"] != "" else None,
                "value_low_pct": row["value_low_pct"] if row["value_low_pct"] != "" else None,
                "comparator": row["comparator"],
                "availability_status": row["availability_status"],
                "indicator_code": row["indicator_code"] or None,
                "display_value": row["display_value"],
                "is_estimated": bool(row["is_estimated"]),
                "featured": True,
                "source_id": row["source_id"],
                "source_row_number": row["source_row_number"],
                "notes": indicator_notes.get(row["v3_mineral"], ""),
            }
            for row in critical_reliance_rows
        ],
        "qualitative_indicators": [
            {
                "id": "rare_earth_mineral_concentrates",
                "label": "Rare-earth mineral concentrates",
                "scope": "Mineral concentrates",
                "display_value": "Net exporter",
                "availability_status": "indicator",
                "indicator_code": "net_exporter",
                "source_id": "usgs_mcs2026_v1_3",
                "source_row_number": 6065,
            }
        ],
        "source_notes": [
            "Net import reliance is U.S. dependence on all foreign sources; it is not a China share or mine-origin measure.",
            "Every displayed 2025 reliance indicator is an MCS estimate, but product scopes and denominators differ by chapter.",
            "The >75% bauxite and >50% tungsten observations are lower bounds, not point estimates.",
            "Rare earths, heavy rare earths, scandium, and yttrium overlap in scope and must not be summed or averaged.",
            "Cross-mineral production is excluded because MCS units, processing stages, and content bases are not comparable.",
        ],
        "download": f"data/processed/{OUTPUT_FILENAMES['critical_reliance']}",
    }

    heavy_reliance_2025 = next(
        row["value"]
        for row in mcs_rows
        if row["mcs_chapter"] == "RARE EARTHS (Heavy)"
        and row["statistics"] == "Net import reliance"
        and row["year"] == 2025
    )
    site_series = [
        {
            "year": row["year"],
            "china_production": row["china_production"],
            "us_production": row["us_production"],
            "world_production": row["world_production"],
            "china_share_of_world_production": row["china_share_percent"],
            "us_share_of_world_production": row["us_share_percent"],
            "source_vintage": (
                "USGS MYB 2022 T8"
                if row["year"] <= 2022
                else (None if row["year"] == 2023 else "USGS MCS 2026 v1.3")
            ),
        }
        for row in share_series
    ]
    latest_contract = {
        "year": 2025,
        "world_production": latest["world_production"],
        "china_production": latest["china_production"],
        "us_production": latest["us_production"],
        "china_share_of_world_production": latest["china_share_percent"],
        "us_share_of_world_production": latest["us_share_percent"],
        "china_to_us_production_ratio": round(
            latest["china_production"] / latest["us_production"], 2
        ),
        # Explicit unit-bearing aliases make the same values safe outside the chart contract.
        "world_production_metric_tons_reo": latest["world_production"],
        "china_production_metric_tons_reo": latest["china_production"],
        "us_production_metric_tons_reo": latest["us_production"],
    }
    rare_snapshot = {
        "period": "2021–24",
        "basis": "direct_or_shipping_source_percent",
        "materials": source_shares("RARE EARTHS"),
        "china_including_hong_kong_percent": 71,
    }
    heavy_snapshot = {
        "period": "2021–24",
        "basis": "direct_or_shipping_source_percent",
        "materials": source_shares("RARE EARTHS (Heavy)"),
        "net_import_reliance_percent_2025": heavy_reliance_2025,
        "not_published": ["Dysprosium", "Thulium"],
    }
    yttrium_snapshot = {
        "period": "2021–24",
        "basis": "direct_or_shipping_source_percent",
        "materials": source_shares("YTTRIUM"),
        "china_including_hong_kong_percent": 70,
    }
    source_notes = [
        "Mine production is upstream supply context, not a direct measure of access, ownership, or control.",
        "Import-source percentages identify direct or shipping sources and may not identify mine origin.",
        "Heavy-rare-earth and yttrium shares exclude material in value-added intermediates and finished goods.",
        "No comparable 2023 world mine-production table is ingested; 2023 is explicitly missing.",
    ]
    return {
        # Stable site contract.
        "status": "loaded",
        "coverage": [2018, 2025],
        "observation_gap": [2023],
        "unit": "metric_tons_reo_equivalent",
        "series": site_series,
        "latest": latest_contract,
        "import_source_snapshot": {
            "rare_earth_compounds_metals_china_share": 71,
            "rare_earth_compounds_metals_period": "2021–24",
            "rare_earth_compounds_metals_scope": (
                "Direct or shipping source for U.S. rare-earth compound and metal imports; "
                "China includes Hong Kong."
            ),
            "heavy_net_import_reliance_2025": heavy_reliance_2025,
            "yttrium_china_direct_share": 70,
            "caveat": (
                "Direct or shipping source is not necessarily mine origin; value-added "
                "intermediates and finished goods are outside several chapter source-share scopes."
            ),
            "rare": rare_snapshot,
            "heavy": heavy_snapshot,
            "yttrium": yttrium_snapshot,
            # Descriptive aliases retained for consumers that avoid abbreviations.
            "rare_earths": rare_snapshot,
            "heavy_rare_earths": heavy_snapshot,
        },
        "source_notes": source_notes,
        "us_statistical_baseline": us_statistical_baseline,
        "us_official_baseline": us_official_baseline,
        # Richer audit-oriented context retained for data downloads and tests.
        "source_ids": ["usgs_myb2022_t8", "usgs_mcs2026_v1_3"],
        "mine_production_share_series": share_series,
        "latest_2025": latest_contract,
        "import_source_snapshot_2021_2024": import_sources,
        "missing_import_source_materials": ["Dysprosium", "Thulium"],
        "interpretation_note": (
            "Mine production and direct import-source shares are context layers, not direct measures "
            "of PRC access, ownership, or control."
        ),
    }


def build_data_dictionary() -> list[dict[str, str]]:
    """Return field definitions for the two normalized CSV outputs."""

    shared_mcs = {
        "raw_value": "Exact value string in the frozen ScienceBase CSV.",
        "raw_notes": "Exact Notes field in the frozen ScienceBase CSV.",
        "current_value": "Value string after only the audited current-PDF revision, if any.",
        "current_notes": "Notes after only the audited current-PDF revision, if any.",
        "value": "Parsed exact numeric value; blank for a range, bound, indicator, or missing value.",
        "value_low": "Published low endpoint or strict lower-bound value; no midpoint is inferred.",
        "value_high": "Published high endpoint for a range; no midpoint is inferred.",
        "comparator": "Publication relationship: exact, greater_than, range, or blank.",
        "availability_status": "available, explicit_zero, indicator, or not_available.",
        "indicator_code": "Semantic code for a nonnumeric publication indicator; E means net_exporter.",
        "revision_action": "Audited action applied to current fields; none leaves raw and current identical.",
    }
    definitions: list[dict[str, str]] = []
    for column in MCS_COLUMNS:
        definition = shared_mcs.get(column, f"Normalized MCS source field: {column.replace('_', ' ')}.")
        definitions.append({"dataset": "usgs_mcs2026_observations", "column": column, "definition": definition})
    myb_specific = {
        "raw_value": "Unformatted cell value from T8.",
        "display_value": "T8 cell value rendered with the source number format; zero remains --.",
        "raw_marker": "Adjacent source marker: e estimated, r revised, 6 footnote 6, or blank.",
        "is_estimated": "True when an e marker applies at row or cell level.",
        "is_revised": "True when an r marker applies to the cell.",
        "source_cell": "Excel coordinate of the T8 value cell.",
        "source_marker_cell": "Excel coordinate of the adjacent annotation cell.",
        "data_status": "reported, estimated, revised, estimated_revised, or explicit_zero.",
    }
    for column in MYB_COLUMNS:
        definition = myb_specific.get(column, f"Normalized MYB T8 field: {column.replace('_', ' ')}.")
        definitions.append(
            {"dataset": "usgs_myb2022_world_mine_production", "column": column, "definition": definition}
        )
    reliance_specific = {
        "v3_mineral": "V3 analytical family receiving this explicitly scoped MCS chapter indicator.",
        "scope": "Material or product scope that qualifies the published reliance percentage.",
        "raw_value": "Exact 2025 value token in the frozen MCS ScienceBase CSV.",
        "display_value": "Public display token with the percent sign and any source comparator preserved.",
        "value_pct": "Exact percentage when the source publishes a point value.",
        "value_low_pct": "Published lower bound for a greater-than observation; never a point estimate.",
        "mapping_note": "Boundary between the MCS chapter measure and the related V3 analytical family.",
    }
    for column in CRITICAL_RELIANCE_COLUMNS:
        definition = reliance_specific.get(
            column,
            f"Normalized MCS critical-mineral reliance field: {column.replace('_', ' ')}.",
        )
        definitions.append(
            {
                "dataset": "usgs_mcs2026_critical_mineral_reliance",
                "column": column,
                "definition": definition,
            }
        )
    return definitions


def _write_csv(path: Path, columns: list[str], rows: Iterable[Mapping[str, Any]]) -> None:
    """Write a stable UTF-8 CSV with LF newlines and the declared column order."""

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n", extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)


def write_usgs_publication_outputs(
    output_dir: Path, raw_dir: Path = RAW_DIR
) -> dict[str, dict[str, Any]]:
    """Build all USGS publication outputs and return metadata and site context.

    This is the module's only output-writing entry point.  Existing directories
    are accepted; the requested output directory is created when needed.
    """

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    mcs_rows = parse_mcs_2026(Path(raw_dir))
    myb_rows = parse_myb_2022_t8(Path(raw_dir))
    critical_reliance_rows = parse_mcs_2026_critical_reliance(Path(raw_dir))
    metadata = build_usgs_publications_metadata(
        mcs_rows,
        myb_rows,
        critical_reliance_rows,
        Path(raw_dir),
    )
    context = build_usgs_publications_context(mcs_rows, myb_rows, critical_reliance_rows)
    revision_rows = [row for row in mcs_rows if row["revision_action"] != "none"]

    _write_csv(output_path / OUTPUT_FILENAMES["mcs_observations"], MCS_COLUMNS, mcs_rows)
    _write_csv(
        output_path / OUTPUT_FILENAMES["mcs_revision_audit"],
        MCS_REVISION_COLUMNS,
        ({column: row[column] for column in MCS_REVISION_COLUMNS} for row in revision_rows),
    )
    _write_csv(output_path / OUTPUT_FILENAMES["myb_world_production"], MYB_COLUMNS, myb_rows)
    _write_csv(
        output_path / OUTPUT_FILENAMES["critical_reliance"],
        CRITICAL_RELIANCE_COLUMNS,
        critical_reliance_rows,
    )
    _write_csv(
        output_path / OUTPUT_FILENAMES["data_dictionary"],
        DATA_DICTIONARY_COLUMNS,
        build_data_dictionary(),
    )
    (output_path / OUTPUT_FILENAMES["metadata"]).write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return {"metadata": metadata, "context": context}


def main() -> None:
    """Command-line wrapper around :func:`write_usgs_publication_outputs`."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--raw-dir",
        type=Path,
        default=RAW_DIR,
        help="Directory containing the frozen USGS inputs (default: data/raw).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        required=True,
        help="Directory in which to write normalized CSV and metadata outputs.",
    )
    args = parser.parse_args()
    result = write_usgs_publication_outputs(args.output_dir, args.raw_dir)
    print(
        json.dumps(
            {
                "mcs_rows": result["metadata"]["datasets"]["mcs2026"]["row_count"],
                "myb_rows": result["metadata"]["datasets"]["myb2022_t8"]["row_count"],
                "revision_rows": result["metadata"]["datasets"]["mcs2026"]["revision_count"],
                "critical_reliance_rows": result["metadata"]["datasets"]["critical_mineral_reliance_2025"]["row_count"],
                "output_dir": str(args.output_dir),
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
