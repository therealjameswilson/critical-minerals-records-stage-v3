#!/usr/bin/env python3
"""Normalize official USITC DataWeb XLSX exports for the V3 public site.

The input workbooks are preserved byte-for-byte under ``data/source-files``.
This script reads the value sheets only. It produces one compact value series
per U.S. trade flow, reported partner, and HTS4 category. Rows split by source
quantity description are summed because DataWeb's HTS4 value report allocates
the four-digit category across those source rows. Quantity measures are not
combined here because their units are not consistently additive.

Interpretation is deliberately narrow:

* Imports for consumption are a U.S.-reported access signal. The partner is
  the source-reported country of origin, not mine origin or ownership.
* Domestic exports are a U.S.-reported outbound bilateral signal. They do not
  measure the destination country's total imports or resource access.
* HTS4 categories are traded-product groupings, not deposits or mineral
  ownership. Broad and mixed categories are labeled as such.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import warnings
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_IMPORTS = ROOT / "data" / "source-files" / "usitc-dataweb-us-imports-1993-2026.xlsx"
DEFAULT_EXPORTS = ROOT / "data" / "source-files" / "usitc-dataweb-us-domestic-exports-1993-2026.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "v3" / "dataweb-series.json"
DEFAULT_REGISTRY = ROOT / "data" / "v3" / "dataset-registry.json"
YEARS = list(range(1993, 2027))
SOURCE_URL = "https://dataweb.usitc.gov/"

# DataWeb's generated XLSX files intentionally omit a workbook default style.
# That does not affect values and otherwise produces a noisy openpyxl warning.
warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


PARTNERS = {
    "Australia": ("aus", "AUS"),
    "Brazil": ("bra", "BRA"),
    "Canada": ("can", "CAN"),
    "Chile": ("chl", "CHL"),
    "China": ("chn", "CHN"),
    "Democratic Republic of the Congo": ("cod", "COD"),
    "Estonia": ("est", "EST"),
    "France": ("fra", "FRA"),
    "Germany": ("deu", "DEU"),
    "Hong Kong": ("hkg", "HKG"),
    "India": ("ind", "IND"),
    "Japan": ("jpn", "JPN"),
    "Malaysia": ("mys", "MYS"),
    "Myanmar (Burma)": ("mmr", "MMR"),
    "Russia": ("rus", "RUS"),
    "South Africa": ("zaf", "ZAF"),
    "South Korea": ("kor", "KOR"),
    "Vietnam": ("vnm", "VNM"),
}


MATERIAL_OVERRIDES = {
    "2504": ("Natural graphite", "Battery and industrial inputs", "named material", "raw material"),
    "2601": ("Iron ores and concentrates", "Bulk strategic materials", "named ore category", "mining feedstock"),
    "2602": ("Manganese ores and concentrates", "Battery and industrial inputs", "named ore category", "mining feedstock"),
    "2603": ("Copper ores and concentrates", "Battery and industrial inputs", "named ore category", "mining feedstock"),
    "2605": ("Cobalt ores and concentrates", "Battery and industrial inputs", "named ore category", "mining feedstock"),
    "2606": ("Aluminum ores and concentrates", "Bulk strategic materials", "named ore category", "mining feedstock"),
    "2609": ("Tin ores and concentrates", "Strategic metals", "named ore category", "mining feedstock"),
    "2610": ("Chromium ores and concentrates", "Strategic metals", "named ore category", "mining feedstock"),
    "2611": ("Tungsten ores and concentrates", "Strategic metals", "named ore category", "mining feedstock"),
    "2612": ("Uranium or thorium ores and concentrates", "Nuclear materials", "combined ore category", "mining feedstock"),
    "2805": ("Rare-earth metals, scandium and yttrium; mercury", "Rare earths and magnets", "mixed named category", "refined material"),
    "2825": ("Selected inorganic bases and metal oxides", "Battery and industrial inputs", "broad mixed category", "processed material"),
    "2836": ("Carbonates and peroxocarbonates", "Battery and industrial inputs", "broad mixed category", "processed material"),
    "2846": ("Rare-earth compounds, yttrium and scandium", "Rare earths and magnets", "named compound category", "processed material"),
    "7202": ("Ferroalloys", "Strategic metals", "broad mixed category", "processed material"),
    "7502": ("Nickel, unwrought", "Battery and industrial inputs", "named material", "refined material"),
    "7901": ("Zinc, unwrought", "Strategic metals", "named material", "refined material"),
    "8001": ("Tin, unwrought", "Strategic metals", "named material", "refined material"),
    "8101": ("Tungsten and articles", "Strategic metals", "named material and articles", "processed material"),
    "8103": ("Tantalum and articles", "Strategic metals", "named material and articles", "processed material"),
    "8105": ("Cobalt products and articles", "Battery and industrial inputs", "named material and articles", "processed material"),
    "8106": ("Bismuth and articles", "Strategic metals", "named material and articles", "processed material"),
    "8110": ("Antimony and articles", "Strategic metals", "named material and articles", "processed material"),
    "8112": ("Selected minor metals and articles", "Strategic metals", "broad mixed category", "processed material"),
    "8505": ("Permanent magnets, electromagnets and parts", "Rare earths and magnets", "manufactured product category", "manufactured product"),
}


FLOW_CONFIG = {
    "imports_for_consumption": {
        "file_role": "imports",
        "sheet": "Customs Value",
        "ytd_sheet": "Customs Value YTD",
        "value_basis": "customs_value_usd",
        "source_flow_label": "Imports For Consumption",
        "partner_role": "reported_origin",
        "access_interpretation": "U.S.-reported import access signal",
    },
    "domestic_exports": {
        "file_role": "exports",
        "sheet": "FAS Value",
        "ytd_sheet": "FAS Value YTD",
        "value_basis": "fas_value_usd",
        "source_flow_label": "Domestic Exports",
        "partner_role": "reported_destination",
        "access_interpretation": "U.S.-reported bilateral outbound signal; not PRC total access",
    },
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def json_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text or text in {"--", "N/A", "NA"}:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def truthy(value: Any) -> bool:
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y", "s", "suppressed"}


def excel_date(value: Any, epoch: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value, epoch).date().isoformat()
    return str(value or "").strip()


def query_parameters(workbook) -> dict[str, Any]:
    sheet = workbook["Query Parameters"]
    # DataWeb exports declare an incorrect A1-only worksheet dimension.  Reset
    # it before streaming so openpyxl reads the complete XML payload.
    sheet.reset_dimensions()
    result: dict[str, Any] = {}
    for row in sheet.iter_rows(min_row=1, max_row=100, max_col=2, values_only=True):
        key, value = row
        if key and value not in (None, ""):
            result[str(key).strip()] = value
    if "Download Date" in result:
        result["Download Date"] = excel_date(result["Download Date"], workbook.epoch)
    return result


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[3], start=1)
        if cell.value not in (None, "")
    }


def value_column(headers: dict[str, int], year: int, *, ytd: bool) -> int:
    candidates = [f"{year}_Year_to_date", f"{year}_YTD"] if ytd else [str(year), year]
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
        if str(candidate) in headers:
            return headers[str(candidate)]
    raise ValueError(f"Missing {'YTD ' if ytd else ''}value column for {year}")


def suppression_column(headers: dict[str, int], year: int, *, ytd: bool) -> int | None:
    names = [f"{year}_Year_to_date_Suppressed", f"{year}_YTD_Suppressed"] if ytd else [f"{year}_Suppressed"]
    return next((headers[name] for name in names if name in headers), None)


def material_record(code: str, description: str) -> dict[str, str]:
    label, group, proxy_type, supply_chain_stage = MATERIAL_OVERRIDES.get(
        code,
        (description.title(), "Other strategic resources", "source-defined HTS4 category", "unspecified traded product"),
    )
    return {
        "hts4": code,
        "label": label,
        "source_description": description,
        "group": group,
        "proxy_type": proxy_type,
        "supply_chain_stage": supply_chain_stage,
    }


def parse_value_sheet(sheet, *, ytd: bool) -> tuple[dict[tuple[str, str], dict[str, Any]], int]:
    sheet.reset_dimensions()
    headers = header_map(sheet)
    required = {"Country", "HTS Number", "Description"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"{sheet.title}: missing headers {sorted(missing)}")
    value_columns = {year: value_column(headers, year, ytd=ytd) for year in YEARS}
    suppression_columns = {year: suppression_column(headers, year, ytd=ytd) for year in YEARS}
    result: dict[tuple[str, str], dict[str, Any]] = {}
    source_rows = 0

    for row in sheet.iter_rows(min_row=4, values_only=True):
        partner = str(row[headers["Country"] - 1] or "").strip()
        raw_code = row[headers["HTS Number"] - 1]
        description = str(row[headers["Description"] - 1] or "").strip()
        if not partner or raw_code in (None, ""):
            continue
        code = str(raw_code).split(".")[0].zfill(4)
        key = (partner, code)
        entry = result.setdefault(
            key,
            {
                "partner": partner,
                "hts4": code,
                "descriptions": set(),
                "values": [None] * len(YEARS),
                "suppressed": [False] * len(YEARS),
                "source_row_count": 0,
            },
        )
        if description:
            entry["descriptions"].add(description)
        entry["source_row_count"] += 1
        source_rows += 1

        for index, year in enumerate(YEARS):
            value = json_number(row[value_columns[year] - 1])
            suppressed_col = suppression_columns[year]
            is_suppressed = truthy(row[suppressed_col - 1]) if suppressed_col else False
            entry["suppressed"][index] = entry["suppressed"][index] or is_suppressed
            if value is not None:
                entry["values"][index] = (entry["values"][index] or 0) + value

    return result, source_rows


def normalize_workbook(path: Path, flow: str) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, dict[str, str]]]:
    config = FLOW_CONFIG[flow]
    workbook = load_workbook(path, read_only=True, data_only=True)
    params = query_parameters(workbook)
    if params.get("Trade Flow") != config["source_flow_label"]:
        raise ValueError(f"{path.name}: expected {config['source_flow_label']}, found {params.get('Trade Flow')}")
    if params.get("Years") != ", ".join(str(year) for year in YEARS):
        raise ValueError(f"{path.name}: query years do not exactly match 1993-2026")
    annual, annual_rows = parse_value_sheet(workbook[config["sheet"]], ytd=False)
    ytd, ytd_rows = parse_value_sheet(workbook[config["ytd_sheet"]], ytd=True)
    materials: dict[str, dict[str, str]] = {}
    series: list[dict[str, Any]] = []
    for partner, code in sorted(set(annual) | set(ytd)):
        if partner not in PARTNERS:
            raise ValueError(f"{path.name}: unknown partner {partner!r}")
        empty_entry = {
            "partner": partner,
            "hts4": code,
            "descriptions": set(),
            "values": [None] * len(YEARS),
            "suppressed": [False] * len(YEARS),
            "source_row_count": 0,
        }
        annual_entry = annual.get((partner, code), empty_entry)
        ytd_entry = ytd.get((partner, code), empty_entry)
        descriptions = sorted(annual_entry["descriptions"] | ytd_entry["descriptions"])
        if not descriptions:
            raise ValueError(f"{path.name}: {partner}/{code} has no description")
        description = descriptions[0]
        materials.setdefault(code, material_record(code, description))
        partner_id, partner_iso3 = PARTNERS[partner]
        # DataWeb leaves the current-year column on the main annual sheet as a
        # zero placeholder; the companion YTD sheet contains the actual
        # January-April value. Build the mixed display series explicitly from
        # full years through 2025 plus the current YTD cell for 2026.
        annual_or_current_ytd = [*annual_entry["values"][:-1], ytd_entry["values"][-1]]
        annual_or_current_ytd_suppressed = [
            *annual_entry["suppressed"][:-1],
            ytd_entry["suppressed"][-1],
        ]
        series.append(
            {
                "id": f"{flow}-{partner_id}-{code}",
                "flow": flow,
                "value_basis": config["value_basis"],
                "focal_country_id": "usa",
                "focal_country_role": "reporting_economy",
                "reporting_economy_id": "usa",
                "reporting_economy_iso3": "USA",
                "partner_country_id": partner_id,
                "partner_iso3": partner_iso3,
                "partner_name": partner,
                "partner_role": config["partner_role"],
                "hts4": code,
                "commodity_description": description,
                "access_interpretation": config["access_interpretation"],
                "annual_or_current_ytd": annual_or_current_ytd,
                "comparable_jan_apr_ytd": ytd_entry["values"],
                "suppressed": {
                    "annual_or_current_ytd": annual_or_current_ytd_suppressed,
                    "comparable_jan_apr_ytd": ytd_entry["suppressed"],
                },
                "aggregation_note": "Summed across DataWeb value rows split by source quantity description within the same partner and HTS4 category.",
                "source_row_count": annual_entry["source_row_count"],
                "source_id": f"usitc-dataweb-{config['file_role']}-2026-07-10",
            }
        )

    manifest = {
        "source_id": f"usitc-dataweb-{config['file_role']}-2026-07-10",
        "title": f"USITC DataWeb export: {config['source_flow_label']}, selected partners and HTS4 categories, 1993-2026",
        "agency": "U.S. International Trade Commission",
        "origin_agency": "U.S. Department of Commerce, Census Bureau",
        "source_url": SOURCE_URL,
        "local_file": str(path.relative_to(ROOT)),
        "sha256": sha256(path),
        "download_date": params["Download Date"],
        "flow": flow,
        "value_basis": config["value_basis"],
        "source_row_count_annual": annual_rows,
        "source_row_count_ytd": ytd_rows,
        "normalized_series_count": len(series),
        "query_parameters": {key: str(value) for key, value in params.items()},
    }
    return series, manifest, materials


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--imports", type=Path, default=DEFAULT_IMPORTS)
    parser.add_argument("--exports", type=Path, default=DEFAULT_EXPORTS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    args = parser.parse_args()

    all_series: list[dict[str, Any]] = []
    manifests: list[dict[str, Any]] = []
    materials: dict[str, dict[str, str]] = {}
    for path, flow in (
        (args.imports.resolve(), "imports_for_consumption"),
        (args.exports.resolve(), "domestic_exports"),
    ):
        series, manifest, workbook_materials = normalize_workbook(path, flow)
        all_series.extend(series)
        manifests.append(manifest)
        for code, material in workbook_materials.items():
            if code in materials and materials[code]["source_description"] != material["source_description"]:
                materials[code]["source_description"] += " / " + material["source_description"]
            else:
                materials.setdefault(code, material)

    partner_records = [
        {"id": partner_id, "iso3": iso3, "name": name, "prc_scope": name == "China"}
        for name, (partner_id, iso3) in PARTNERS.items()
    ]
    output = {
        "schema_version": "3.0.0",
        "generated_at": f"{max(manifest['download_date'] for manifest in manifests)}T00:00:00Z",
        "coverage": {
            "start_year": 1993,
            "end_year": 2026,
            "current_ytd_months": "January-April",
            "reporting_economy": "United States",
            "reporting_economy_iso3": "USA",
            "selected_partner_count": len(PARTNERS),
            "commodity_count": len(materials),
            "series_count": len(all_series),
        },
        "periods": {
            "annual_or_current_ytd": {
                "label": "Full years through 2025; January-April YTD for 2026",
                "comparable_across_all_years": False,
            },
            "comparable_jan_apr_ytd": {
                "label": "January-April year-to-date for every year",
                "comparable_across_all_years": True,
            },
        },
        "years": YEARS,
        "partners": partner_records,
        "materials": [materials[code] for code in sorted(materials)],
        "series": sorted(all_series, key=lambda row: row["id"]),
    }
    registry = {
        "schema_version": "3.0.0",
        "updated": "2026-07-10",
        "method": "Official U.S. Government statistics only",
        "comparison_status": {
            "us_access": "live_partial",
            "prc_access": "awaiting_prc_reporter_or_comparable_usg_series",
            "bilateral_us_prc": "live_us_reporter_only",
        },
        "sources": manifests,
        "known_limits": [
            "All loaded trade series use the United States as the reporting economy.",
            "A U.S. import partner is a reported origin, not necessarily mine origin, ownership, processing control, route, or end use.",
            "A U.S. export destination does not measure that economy's total imports or strategic access.",
            "HTS4 categories are traded-product groupings; several combine multiple minerals or manufactured products.",
            "The 2026 annual-or-current-YTD column covers January-April, so it must not be compared directly with full prior years.",
            "Hong Kong is preserved as a separate source-reported partner and is not combined with China.",
        ],
        "planned_comparison_layers": [
            {"id": "trade", "label": "Trade access", "us_status": "live_partial", "prc_status": "not_loaded"},
            {"id": "mine-production", "label": "Mine production", "us_status": "not_loaded", "prc_status": "not_loaded"},
            {"id": "processing", "label": "Processing and refining", "us_status": "not_loaded", "prc_status": "not_loaded"},
            {"id": "reserves", "label": "Reserves and resources", "us_status": "not_loaded", "prc_status": "not_loaded"},
            {"id": "stockpiles", "label": "Government stockpiles", "us_status": "not_loaded", "prc_status": "not_loaded"},
        ],
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    args.registry.write_text(json.dumps(registry, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(all_series)} normalized value series to {args.output}")
    print(f"Wrote {len(manifests)} source manifests to {args.registry}")


if __name__ == "__main__":
    main()
