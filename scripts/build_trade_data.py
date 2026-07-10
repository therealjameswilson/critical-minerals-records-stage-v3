#!/usr/bin/env python3
"""Build the V3 statistical record from frozen DataWeb workbooks.

The workbooks are U.S.-reporter extracts for 18 selected partners.  They do
not contain a World row.  Consequently, every U.S. share produced here uses
the explicit denominator ``selected_18_partners``.

Value and first-quantity rows share a lossless source key.  Second quantities
do not identify their corresponding first-unit value bucket at HTS4, so they
are emitted as separate measure rows rather than duplicated across buckets.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
PROCESSED = ROOT / "data" / "processed"
EXPLORER = PROCESSED / "explorer"
YEARS = tuple(range(1993, 2027))
RETRIEVED_AT = "2026-07-10"
SOURCE_URL = "https://dataweb.usitc.gov/"
DENOMINATOR_SCOPE = "selected_18_partners"
RARE_EARTH_HTS = {"2805", "2846", "8505"}
USGS_DS140_FILENAME = "usgs_ds140_rare_earths_2020.xlsx"
USGS_DS140_PAGE_URL = "https://www.usgs.gov/media/files/rare-earths-historical-statistics-data-series-140"
USGS_DS140_DOWNLOAD_URL = "https://d9-wret.s3.us-west-2.amazonaws.com/assets/palladium/production/s3fs-public/media/files/ds140-rare-earths-2020.xlsx"
USGS_DS140_LAST_MODIFIED = "2023-09-27"
USGS_DS140_PACKAGE_MODIFIED_AT = "2024-06-04T18:22:38Z"
USGS_DS140_PUBLICATION_DATE = "2024-06-04"
USGS_DS140_EMBEDDED_NOTES_PATH = "xl/embeddings/Microsoft_Word_Document.docx"

warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


@dataclass(frozen=True)
class FlowConfig:
    flow: str
    filename: str
    annual_value_sheet: str
    ytd_value_sheet: str
    value_basis: str
    partner_role: str


FLOWS = (
    FlowConfig(
        "imports_for_consumption",
        "us_imports_for_consumption_1993-2026.xlsx",
        "Customs Value",
        "Customs Value YTD",
        "customs_value_usd",
        "country_of_origin",
    ),
    FlowConfig(
        "domestic_exports",
        "us_domestic_exports_1993-2026.xlsx",
        "FAS Value",
        "FAS Value YTD",
        "fas_value_usd",
        "ultimate_destination",
    ),
)


PARTNER_ISO = {
    "Australia": "AUS",
    "Brazil": "BRA",
    "Canada": "CAN",
    "Chile": "CHL",
    "China": "CHN",
    "Democratic Republic of the Congo": "COD",
    "Estonia": "EST",
    "France": "FRA",
    "Germany": "DEU",
    "Hong Kong": "HKG",
    "India": "IND",
    "Japan": "JPN",
    "Malaysia": "MYS",
    "Myanmar (Burma)": "MMR",
    "Russia": "RUS",
    "South Africa": "ZAF",
    "South Korea": "KOR",
    "Vietnam": "VNM",
}


# label, mineral, processing stage, scope note
MATERIALS = {
    "2504": ("Natural graphite", "natural_graphite", "ore", "Natural graphite in primary forms."),
    "2601": ("Iron ores and concentrates", "iron", "ore", "Ore and concentrate proxy."),
    "2602": ("Manganese ores and concentrates", "manganese", "ore", "Ore and concentrate proxy."),
    "2603": ("Copper ores and concentrates", "copper", "ore", "Ore and concentrate proxy."),
    "2605": ("Cobalt ores and concentrates", "cobalt", "ore", "Ore and concentrate proxy."),
    "2606": ("Aluminum ores and concentrates", "aluminum", "ore", "Bauxite and other aluminum ores."),
    "2609": ("Tin ores and concentrates", "tin", "ore", "Ore and concentrate proxy."),
    "2610": ("Chromium ores and concentrates", "chromium", "ore", "Ore and concentrate proxy."),
    "2611": ("Tungsten ores and concentrates", "tungsten", "ore", "Ore and concentrate proxy."),
    "2612": ("Uranium or thorium ores", "uranium_thorium", "ore", "Combined uranium and thorium category."),
    "2805": ("Rare-earth metals and related metals", "rare_earths", "metal", "Broad HTS4 proxy; also includes alkali metals and mercury."),
    "2825": ("Selected metal oxides and hydroxides", "mixed_metal_oxides", "compound", "Broad mixed chemical category."),
    "2836": ("Carbonates and peroxocarbonates", "mixed_carbonates", "compound", "Broad mixed chemical category."),
    "2846": ("Rare-earth compounds", "rare_earths", "compound", "Compounds of rare-earth metals, yttrium, or scandium."),
    "7202": ("Ferroalloys", "ferroalloys", "alloy", "Broad mixed alloy category."),
    "7502": ("Nickel, unwrought", "nickel", "metal", "Unwrought nickel."),
    "7901": ("Zinc, unwrought", "zinc", "metal", "Unwrought zinc."),
    "8001": ("Tin, unwrought", "tin", "metal", "Unwrought tin."),
    "8101": ("Tungsten and articles", "tungsten", "metal", "Tungsten plus downstream articles."),
    "8103": ("Tantalum and articles", "tantalum", "metal", "Tantalum plus downstream articles."),
    "8105": ("Cobalt products and articles", "cobalt", "metal", "Cobalt mattes and other products/articles."),
    "8106": ("Bismuth and articles", "bismuth", "metal", "Bismuth plus downstream articles."),
    "8110": ("Antimony and articles", "antimony", "metal", "Antimony plus downstream articles."),
    "8112": ("Selected minor metals and articles", "minor_metals", "metal", "Broad mixed minor-metals category."),
    "8505": ("Magnets, electromagnets, and parts", "rare_earths", "magnet", "Broad magnet proxy; includes non-rare-earth and non-permanent-magnet products."),
}


# Source heading -> metric id, geography, analytical code, normalized unit, scope note, method-note id.
USGS_DS140_METRICS = {
    "Production": (
        "production",
        "United States",
        "USA",
        "metric_tons_reo_equivalent",
        "Contained rare-earth-oxide equivalent in U.S.-produced bastnaesite and monazite concentrates.",
        "embedded-notes:production",
    ),
    "Imports": (
        "imports",
        "United States",
        "USA",
        "metric_tons_reo_equivalent",
        "Estimated rare-earth-oxide equivalent in alloys, compounds, metals, and ores imported into the United States.",
        "embedded-notes:imports",
    ),
    "Exports": (
        "exports",
        "United States",
        "USA",
        "metric_tons_reo_equivalent",
        "Estimated rare-earth-oxide equivalent in alloys, compounds, metals, and ores exported from the United States.",
        "embedded-notes:exports",
    ),
    "Apparent consumption": (
        "apparent_consumption",
        "United States",
        "USA",
        "metric_tons_reo_equivalent",
        "USGS apparent consumption; calculated, estimated, or interpolated according to the embedded worksheet notes.",
        "embedded-notes:apparent-consumption",
    ),
    "Unit value ($/t)": (
        "unit_value_current_usd",
        "United States",
        "USA",
        "usd_per_metric_ton_reo_current",
        "Weighted-average U.S. unit value in current dollars per metric ton of apparent consumption.",
        "embedded-notes:unit-value-current",
    ),
    "Unit value (98$/t)": (
        "unit_value_constant_1998_usd",
        "United States",
        "USA",
        "usd_per_metric_ton_reo_constant_1998",
        "U.S. unit value adjusted with the Consumer Price Index to constant 1998 dollars.",
        "embedded-notes:unit-value-constant-1998",
    ),
    "World production": (
        "production",
        "World",
        "WLD",
        "metric_tons_reo_equivalent",
        "World rare-earth-oxide equivalent content of ores produced.",
        "embedded-notes:world-production",
    ),
}


CSV_COLUMNS = [
    "reporter",
    "flow",
    "partner",
    "partner_iso",
    "hts",
    "hts_desc",
    "mineral",
    "processing_stage",
    "year",
    "ytd_flag",
    "value_usd",
    "qty1",
    "qty1_unit",
    "qty2",
    "qty2_unit",
    "source",
    "retrieved_at",
    "period_label",
    "source_quantity_bucket",
    "quantity_measure_slot",
    "suppression_raw",
    "quantity_incomplete",
    "data_status",
    "missing_data_label",
    "value_basis",
    "partner_role",
    "source_workbook",
    "source_sheet",
    "source_row_id",
    "classification_status",
    "break_note_id",
    "denominator_scope",
    "mass_kg",
    "mass_status",
    "scope_note",
]

USGS_DS140_COLUMNS = [
    "source_id",
    "commodity",
    "year",
    "period_type",
    "geography",
    "geography_code",
    "metric",
    "metric_label",
    "value",
    "unit",
    "value_status",
    "method_status",
    "method_note_id",
    "source_value_raw",
    "source_formula",
    "scope_note",
    "source",
    "source_file",
    "source_sheet",
    "source_cell",
    "source_url",
    "download_url",
    "worksheet_last_modified",
    "package_modified_at",
    "source_publication_date",
    "retrieved_at",
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")


def number(value: Any) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        value = float(value)
        if not math.isfinite(value):
            return None
        return int(value) if value.is_integer() else value
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text or text.casefold() in {"n/a", "na", "--"}:
        return None
    parsed = float(text)
    return int(parsed) if parsed.is_integer() else parsed


def normalize_unit(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"^value\s+for:\s*", "", text)
    aliases = {
        "kilograms": "kg",
        "metric tons": "metric_ton",
        "number": "number",
        "no units collected": "not_collected",
        "grams": "gram",
        "gold content grams": "gold_content_gram",
        "component grams": "component_gram",
        "component kilograms": "component_kg",
        "component tons": "component_ton",
    }
    return aliases.get(text, slug(text).replace("-", "_") or "unknown")


def normalize_description(value: Any) -> str:
    return " ".join(str(value or "").split())


def read_query_rows(workbook: Any) -> list[dict[str, str]]:
    sheet = workbook["Query Parameters"]
    sheet.reset_dimensions()
    rows: list[dict[str, str]] = []
    section = ""
    for left, right in sheet.iter_rows(min_row=1, max_row=100, max_col=2, values_only=True):
        if left in (None, "") and right in (None, ""):
            continue
        key = str(left or "").strip()
        if key.startswith("Step "):
            section = key
            continue
        if isinstance(right, datetime):
            rendered = right.isoformat(timespec="seconds")
        elif isinstance(right, date):
            rendered = right.isoformat()
        elif isinstance(right, (int, float)) and key == "Download Date":
            rendered = from_excel(right, workbook.epoch).isoformat(timespec="seconds")
        else:
            rendered = str(right or "").strip()
        rows.append({"section": section, "parameter": key, "value": rendered})
    return rows


USGS_APPARENT_CONSUMPTION_INTERPOLATED_YEARS = {
    *range(1911, 1915),
    *range(1918, 1922),
    1932,
    *range(1942, 1945),
    *range(1946, 1950),
}


def usgs_method_status(header: str, year: int, value_status: str) -> str:
    """Map only the methods explicitly described in the embedded USGS notes."""

    if value_status == "withheld":
        return "withheld_to_avoid_proprietary_disclosure"
    if value_status == "not_available":
        if header == "Apparent consumption" and year == 2011:
            return "negative_calculation_published_as_not_available"
        return "not_available"
    if header in {"Production", "World production"}:
        return "source_series_reo_content_method_not_cell_specific"
    if header in {"Imports", "Exports"}:
        return "estimated_reo_equivalent"
    if header == "Unit value ($/t)":
        return "estimated_weighted_average_imports_exports"
    if header == "Unit value (98$/t)":
        return "calculated_cpi_adjustment_1998_base"
    if header == "Apparent consumption":
        if year in USGS_APPARENT_CONSUMPTION_INTERPOLATED_YEARS:
            return "interpolated"
        if year <= 1999:
            return "estimated_material_balance"
        if year <= 2008:
            return "estimated_material_balance_using_estimated_reo"
        return "calculated_using_estimated_reo"
    raise ValueError(f"unmapped USGS method: {header}/{year}/{value_status}")


def parse_usgs_ds140() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Normalize the USGS Data Series 140 rare-earth worksheet.

    This national REO-equivalent balance is intentionally separate from the
    partner-level HTS trade table. Method labels follow only the series and
    year ranges identified in the embedded source notes; formulas are retained
    exactly and no per-cell reported/estimated label is invented.
    """

    path = RAW / USGS_DS140_FILENAME
    with ZipFile(path) as archive:
        embedded_notes = archive.read(USGS_DS140_EMBEDDED_NOTES_PATH)
        core_properties = archive.read("docProps/core.xml").decode("utf-8")
    embedded_notes_sha256 = hashlib.sha256(embedded_notes).hexdigest()
    package_modified_match = re.search(r"<dcterms:modified[^>]*>([^<]+)</dcterms:modified>", core_properties)
    package_modified_at = package_modified_match.group(1) if package_modified_match else ""
    if package_modified_at != USGS_DS140_PACKAGE_MODIFIED_AT:
        raise ValueError(f"{USGS_DS140_FILENAME}: unexpected package modification timestamp {package_modified_at!r}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    formula_workbook = load_workbook(path, read_only=True, data_only=False)
    if workbook.sheetnames != ["Rare earths"] or formula_workbook.sheetnames != ["Rare earths"]:
        raise ValueError(f"{USGS_DS140_FILENAME}: unexpected sheets {workbook.sheetnames}")
    sheet = workbook["Rare earths"]
    formula_sheet = formula_workbook["Rare earths"]
    sheet.reset_dimensions()
    formula_sheet.reset_dimensions()
    source_rows = list(sheet.iter_rows(min_row=1, max_row=129, min_col=1, max_col=8, values_only=True))
    formula_rows = list(formula_sheet.iter_rows(min_row=1, max_row=129, min_col=1, max_col=8, values_only=True))
    workbook.close()
    formula_workbook.close()

    if normalize_description(source_rows[0][0]) != "RARE EARTHS STATISTICS1":
        raise ValueError(f"{USGS_DS140_FILENAME}: unexpected title")
    unit_note = normalize_description(source_rows[2][0]).strip("[]")
    last_modified_text = normalize_description(source_rows[3][0])
    if last_modified_text != "Last modification: September 27, 2023":
        raise ValueError(f"{USGS_DS140_FILENAME}: unexpected worksheet modification label")
    expected_headers = ["Year", *USGS_DS140_METRICS]
    headers = [normalize_description(value) for value in source_rows[4]]
    if headers != expected_headers:
        raise ValueError(f"{USGS_DS140_FILENAME}: unexpected headers {headers}")

    rows: list[dict[str, Any]] = []
    observed_years: list[int] = []
    status_counts: dict[str, int] = defaultdict(int)
    formula_cells: list[str] = []
    for source_row_number, (source_row, formula_row) in enumerate(zip(source_rows[5:126], formula_rows[5:126]), start=6):
        year_value = number(source_row[0])
        if year_value is None or int(year_value) != year_value:
            raise ValueError(f"{USGS_DS140_FILENAME}: invalid year at row {source_row_number}")
        year = int(year_value)
        observed_years.append(year)
        for column_number, header in enumerate(expected_headers[1:], start=2):
            metric, geography, geography_code, unit, scope_note, method_note_id = USGS_DS140_METRICS[header]
            raw_value = source_row[column_number - 1]
            raw_text = str(raw_value if raw_value is not None else "").strip()
            formula_value = formula_row[column_number - 1]
            source_formula = str(formula_value).strip() if isinstance(formula_value, str) and formula_value.startswith("=") else ""
            source_cell = f"{get_column_letter(column_number)}{source_row_number}"
            if source_formula:
                formula_cells.append(source_cell)
            if raw_text.casefold() == "na" or raw_value in (None, ""):
                value = None
                value_status = "not_available"
            elif raw_text.casefold() == "w":
                value = None
                value_status = "withheld"
            else:
                value = number(raw_value)
                if value is None:
                    raise ValueError(
                        f"{USGS_DS140_FILENAME}: nonnumeric value {raw_value!r} at row {source_row_number}"
                    )
                value_status = "available"
            method_status = usgs_method_status(header, year, value_status)
            status_counts[value_status] += 1
            rows.append(
                {
                    "source_id": "usgs-ds140-rare-earths-2020",
                    "commodity": "rare_earths",
                    "year": year,
                    "period_type": "annual",
                    "geography": geography,
                    "geography_code": geography_code,
                    "metric": metric,
                    "metric_label": header,
                    "value": value,
                    "unit": unit,
                    "value_status": value_status,
                    "method_status": method_status,
                    "method_note_id": method_note_id,
                    "source_value_raw": raw_text,
                    "source_formula": source_formula,
                    "scope_note": scope_note,
                    "source": "U.S. Geological Survey Data Series 140",
                    "source_file": f"data/raw/{USGS_DS140_FILENAME}",
                    "source_sheet": "Rare earths",
                    "source_cell": source_cell,
                    "source_url": USGS_DS140_PAGE_URL,
                    "download_url": USGS_DS140_DOWNLOAD_URL,
                    "worksheet_last_modified": USGS_DS140_LAST_MODIFIED,
                    "package_modified_at": package_modified_at,
                    "source_publication_date": USGS_DS140_PUBLICATION_DATE,
                    "retrieved_at": RETRIEVED_AT,
                }
            )
    if observed_years != list(range(1900, 2021)):
        raise ValueError(f"{USGS_DS140_FILENAME}: expected complete 1900-2020 year rows")
    if formula_cells != [f"E{row}" for row in range(107, 113)]:
        raise ValueError(f"{USGS_DS140_FILENAME}: unexpected formula cells {formula_cells}")

    notes = {
        "data_sources": "U.S. Bureau of Mines and U.S. Geological Survey Minerals Yearbook and Mineral Resources of the United States.",
        "production": USGS_DS140_METRICS["Production"][4],
        "imports": USGS_DS140_METRICS["Imports"][4],
        "exports": USGS_DS140_METRICS["Exports"][4],
        "apparent_consumption": (
            "Estimated for specified years as production plus imports minus exports; interpolated for years listed "
            "in the embedded worksheet notes; 2000 onward calculated using estimated REO content."
        ),
        "unit_value": USGS_DS140_METRICS["Unit value ($/t)"][4],
        "constant_unit_value": USGS_DS140_METRICS["Unit value (98$/t)"][4],
        "world_production": USGS_DS140_METRICS["World production"][4],
        "status_codes": normalize_description(source_rows[126][0]),
        "general": normalize_description(source_rows[128][0]),
    }
    metadata = {
        "source_id": "usgs-ds140-rare-earths-2020",
        "agency": "U.S. Geological Survey",
        "center": "National Minerals Information Center",
        "title": "Rare Earths - Historical Statistics (Data Series 140)",
        "file": f"data/raw/{USGS_DS140_FILENAME}",
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
        "source_page_url": USGS_DS140_PAGE_URL,
        "download_url": USGS_DS140_DOWNLOAD_URL,
        "source_publication_date": USGS_DS140_PUBLICATION_DATE,
        "source_last_modified": USGS_DS140_LAST_MODIFIED,
        "retrieved_at": RETRIEVED_AT,
        "coverage": [1900, 2020],
        "data_year_count": len(observed_years),
        "normalized_rows": len(rows),
        "formula_cells": formula_cells,
        "package_modified_at": package_modified_at,
        "embedded_notes_file": USGS_DS140_EMBEDDED_NOTES_PATH,
        "embedded_notes_sha256": embedded_notes_sha256,
        "status_counts": dict(sorted(status_counts.items())),
        "unit_note": unit_note,
        "compiled_by": normalize_description(source_rows[127][0]),
        "notes": notes,
        "public_domain": True,
        "citation": (
            "U.S. Geological Survey, 2014, Rare earth statistics, in Kelly, T.D., and Matos, G.R., comps., "
            "Historical statistics for mineral and material commodities in the United States: U.S. Geological "
            "Survey Data Series 140, accessed July 10, 2026, at "
            "https://www.usgs.gov/centers/national-minerals-information-center/"
            "historical-statistics-mineral-and-material-commodities."
        ),
        "comparability_warning": (
            "National REO-equivalent statistics are not partner-level HTS trade data and must not enter the "
            "DataWeb China-share, supplier-HHI, or HTS unit-value calculations."
        ),
    }
    return rows, metadata


def build_usgs_site_context(rows: list[dict[str, Any]], metadata: dict[str, Any]) -> dict[str, Any]:
    by_key = {(row["year"], row["geography_code"], row["metric"]): row for row in rows}
    displayed = []
    for year in range(1993, 2021):
        def value(geography_code: str, metric: str) -> int | float | None:
            row = by_key[(year, geography_code, metric)]
            return row["value"] if row["value_status"] == "available" else None

        production = value("USA", "production")
        world_production = value("WLD", "production")
        displayed.append(
            {
                "year": year,
                "us_production": production,
                "us_imports": value("USA", "imports"),
                "us_exports": value("USA", "exports"),
                "us_apparent_consumption": value("USA", "apparent_consumption"),
                "world_production": world_production,
                "us_share_of_world_production": (
                    round(float(production) / float(world_production), 10)
                    if production is not None and world_production not in (None, 0)
                    else None
                ),
            }
        )
    latest = displayed[-1]
    return {
        "status": "loaded",
        "source_id": metadata["source_id"],
        "full_coverage": metadata["coverage"],
        "displayed_coverage": [1993, 2020],
        "unit": "metric_tons_reo_equivalent",
        "source_last_modified": metadata["source_last_modified"],
        "series": displayed,
        "latest": latest,
        "warnings": [
            "This USGS national balance ends in 2020 and is not extended or interpolated.",
            "Production, imports, exports, and apparent consumption are REO-equivalent estimates, not HTS values or partner-origin measures.",
            "The embedded notes identify methods by series and year range; individual production cells are not labeled reported or estimated.",
            "USGS apparent consumption can reflect inventory changes and unattributed trade; the 2011 source value is unavailable after a negative calculation.",
        ],
    }


def header_map(row: Iterable[Any]) -> dict[str, int]:
    return {str(value).strip(): index for index, value in enumerate(row) if value not in (None, "")}


def year_header(year: int, ytd: bool) -> str:
    return f"{year}_Year_to_date" if ytd else str(year)


def suppression_header(year: int, ytd: bool) -> str:
    return f"{year_header(year, ytd)}_Suppressed"


def parse_measure_sheet(workbook: Any, sheet_name: str, *, ytd: bool, measure: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    sheet.reset_dimensions()
    preview = list(sheet.iter_rows(min_row=2, max_row=3, values_only=True))
    source_count = int(number(preview[0][1]) or 0)
    headers = header_map(preview[1])
    required = {"Country", "HTS Number", "Description", "Quantity Description"}
    if not required <= set(headers):
        raise ValueError(f"{sheet_name}: missing headers {sorted(required - set(headers))}")
    for year in YEARS:
        if year_header(year, ytd) not in headers:
            raise ValueError(f"{sheet_name}: missing {year_header(year, ytd)}")

    parsed: list[dict[str, Any]] = []
    for source_index, row in enumerate(
        sheet.iter_rows(min_row=4, max_row=3 + source_count, values_only=True), start=4
    ):
        partner = str(row[headers["Country"]] or "").strip()
        raw_hts = row[headers["HTS Number"]]
        if not partner or raw_hts in (None, ""):
            raise ValueError(f"{sheet_name}: blank key at source row {source_index}")
        hts = str(int(raw_hts)).zfill(4) if isinstance(raw_hts, (int, float)) else str(raw_hts).split(".")[0].zfill(4)
        if partner not in PARTNER_ISO:
            raise ValueError(f"{sheet_name}: unmapped partner {partner!r}")
        if hts not in MATERIALS:
            raise ValueError(f"{sheet_name}: unmapped HTS4 {hts}")
        unit_raw = row[headers["Quantity Description"]]
        unit = normalize_unit(unit_raw)
        values = {year: number(row[headers[year_header(year, ytd)]]) for year in YEARS}
        suppressions = {
            year: int(number(row[headers[suppression_header(year, ytd)]]) or 0)
            if suppression_header(year, ytd) in headers
            else 0
            for year in YEARS
        }
        parsed.append(
            {
                "partner": partner,
                "partner_iso": PARTNER_ISO[partner],
                "hts": hts,
                "description": normalize_description(row[headers["Description"]]),
                "unit": unit,
                "unit_raw": str(unit_raw or "").strip(),
                "values": values,
                "suppressions": suppressions,
                "source_row": source_index,
                "measure": measure,
                "sheet": sheet_name,
            }
        )
    if len(parsed) != source_count:
        raise ValueError(f"{sheet_name}: expected {source_count} source rows; parsed {len(parsed)}")
    return parsed


def source_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return row["partner"], row["hts"], row["description"], row["unit"]


def mass_value(value: int | float | None, unit: str) -> tuple[int | float | None, str]:
    if value is None:
        return None, "missing"
    if unit == "kg":
        return value, "reported_kg"
    if unit == "metric_ton":
        return value * 1000, "converted_metric_ton_to_kg"
    return None, "not_a_total_mass_unit"


def classification_fields(hts: str, year: int) -> tuple[str, str]:
    if hts == "8505" and year == 2019:
        return "measurement_break", "hts8505-unit-regime-2019"
    if year in {1996, 2002, 2007, 2012, 2017, 2022}:
        return "hs_revision_boundary_not_assessed", f"hs-revision-{year}"
    return "continuity_not_assessed", ""


def data_status(value: int | float | None, *, annual_current_year: bool, quantity_only: bool = False) -> tuple[str, str]:
    if annual_current_year:
        return "not_available", "2026 annual total not available; source cell is a structural zero"
    if value is None:
        if quantity_only:
            return "source_blank", "Quantity source cell is blank"
        return "source_blank", "Value source cell is blank"
    if value == 0:
        return "reported_zero", ""
    return "reported", ""


def emit_flow(config: FlowConfig) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = RAW / config.filename
    workbook = load_workbook(path, read_only=True, data_only=True)
    query_rows = read_query_rows(workbook)
    emitted: list[dict[str, Any]] = []
    row_inventory: dict[str, int] = {}

    for ytd, value_sheet, first_sheet, second_sheet in (
        (False, config.annual_value_sheet, "First Unit of Quantity", "Second Unit of Quantity"),
        (True, config.ytd_value_sheet, "First Unit of Quantity YTD", "Second Unit of Quantity YTD"),
    ):
        values = parse_measure_sheet(workbook, value_sheet, ytd=ytd, measure="value")
        first = parse_measure_sheet(workbook, first_sheet, ytd=ytd, measure="first")
        second = parse_measure_sheet(workbook, second_sheet, ytd=ytd, measure="second")
        row_inventory[value_sheet] = len(values)
        row_inventory[first_sheet] = len(first)
        row_inventory[second_sheet] = len(second)

        value_by_key = {source_key(row): row for row in values}
        first_by_key = {source_key(row): row for row in first}
        if len(value_by_key) != len(values) or len(first_by_key) != len(first):
            raise ValueError(f"{config.filename}: duplicate value/first source key")
        if set(value_by_key) != set(first_by_key):
            only_value = sorted(set(value_by_key) - set(first_by_key))[:3]
            only_first = sorted(set(first_by_key) - set(value_by_key))[:3]
            raise ValueError(f"{config.filename}: value/Q1 key mismatch: {only_value=} {only_first=}")

        period_label = "January-April YTD" if ytd else "full year"
        for key in sorted(value_by_key):
            value_row = value_by_key[key]
            first_row = first_by_key[key]
            label, mineral, stage, scope_note = MATERIALS[value_row["hts"]]
            for year in YEARS:
                annual_current_year = not ytd and year == 2026
                value = None if annual_current_year else value_row["values"][year]
                qty1 = None if annual_current_year else first_row["values"][year]
                suppression_raw = 0 if annual_current_year else first_row["suppressions"][year]
                incomplete = suppression_raw > 0
                status, missing_label = data_status(value, annual_current_year=annual_current_year)
                mass, mass_status = mass_value(qty1, first_row["unit"])
                if incomplete:
                    mass_status = "incomplete_quantity_suppression"
                classification_status, break_id = classification_fields(value_row["hts"], year)
                emitted.append(
                    {
                        "reporter": "US",
                        "flow": config.flow,
                        "partner": value_row["partner"],
                        "partner_iso": value_row["partner_iso"],
                        "hts": value_row["hts"],
                        "hts_desc": value_row["description"],
                        "mineral": mineral,
                        "processing_stage": stage,
                        "year": year,
                        "ytd_flag": str(ytd).lower(),
                        "value_usd": value,
                        "qty1": qty1,
                        "qty1_unit": first_row["unit"],
                        "qty2": None,
                        "qty2_unit": "",
                        "source": "USITC DataWeb",
                        "retrieved_at": RETRIEVED_AT,
                        "period_label": period_label,
                        "source_quantity_bucket": first_row["unit_raw"],
                        "quantity_measure_slot": "first",
                        "suppression_raw": suppression_raw,
                        "quantity_incomplete": str(incomplete).lower(),
                        "data_status": status,
                        "missing_data_label": missing_label,
                        "value_basis": config.value_basis,
                        "partner_role": config.partner_role,
                        "source_workbook": config.filename,
                        "source_sheet": value_sheet,
                        "source_row_id": f"{config.flow}:{'ytd' if ytd else 'annual'}:{value_row['source_row']}",
                        "classification_status": classification_status,
                        "break_note_id": break_id,
                        "denominator_scope": DENOMINATOR_SCOPE,
                        "mass_kg": mass,
                        "mass_status": mass_status,
                        "scope_note": scope_note,
                    }
                )

        # Q2 remains an independent measure row.  The HTS4 export does not
        # expose a lossless key linking it to a particular Q1/value bucket.
        for second_row in sorted(second, key=source_key):
            label, mineral, stage, scope_note = MATERIALS[second_row["hts"]]
            for year in YEARS:
                annual_current_year = not ytd and year == 2026
                qty2 = None if annual_current_year else second_row["values"][year]
                suppression_raw = 0 if annual_current_year else second_row["suppressions"][year]
                incomplete = suppression_raw > 0
                status, missing_label = data_status(
                    qty2, annual_current_year=annual_current_year, quantity_only=True
                )
                mass, mass_status = mass_value(qty2, second_row["unit"])
                if incomplete:
                    mass_status = "incomplete_quantity_suppression"
                classification_status, break_id = classification_fields(second_row["hts"], year)
                emitted.append(
                    {
                        "reporter": "US",
                        "flow": config.flow,
                        "partner": second_row["partner"],
                        "partner_iso": second_row["partner_iso"],
                        "hts": second_row["hts"],
                        "hts_desc": second_row["description"],
                        "mineral": mineral,
                        "processing_stage": stage,
                        "year": year,
                        "ytd_flag": str(ytd).lower(),
                        "value_usd": None,
                        "qty1": None,
                        "qty1_unit": "",
                        "qty2": qty2,
                        "qty2_unit": second_row["unit"],
                        "source": "USITC DataWeb",
                        "retrieved_at": RETRIEVED_AT,
                        "period_label": period_label,
                        "source_quantity_bucket": second_row["unit_raw"],
                        "quantity_measure_slot": "second",
                        "suppression_raw": suppression_raw,
                        "quantity_incomplete": str(incomplete).lower(),
                        "data_status": status,
                        "missing_data_label": missing_label,
                        "value_basis": "quantity_only_no_lossless_hts4_value_join",
                        "partner_role": config.partner_role,
                        "source_workbook": config.filename,
                        "source_sheet": second_sheet,
                        "source_row_id": f"{config.flow}:{'ytd' if ytd else 'annual'}:q2:{second_row['source_row']}",
                        "classification_status": classification_status,
                        "break_note_id": break_id,
                        "denominator_scope": DENOMINATOR_SCOPE,
                        "mass_kg": mass,
                        "mass_status": mass_status,
                        "scope_note": scope_note,
                    }
                )

    workbook.close()
    metadata = {
        "flow": config.flow,
        "file": f"data/raw/{config.filename}",
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
        "query_parameters": query_rows,
        "source_row_inventory": row_inventory,
        "normalized_rows": len(emitted),
    }
    return emitted, metadata


def write_csv(path: Path, rows: Iterable[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def usable_value_rows(rows: Iterable[dict[str, Any]]) -> Iterable[dict[str, Any]]:
    return (
        row
        for row in rows
        if row["quantity_measure_slot"] == "first" and row["value_usd"] is not None
    )


def build_china_share(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, int, str], dict[str, Any]] = defaultdict(
        lambda: {"china": 0.0, "total": 0.0, "partners": set(), "codes": set()}
    )
    for row in usable_value_rows(rows):
        if row["flow"] != "imports_for_consumption":
            continue
        key = row["mineral"], int(row["year"]), row["ytd_flag"]
        value = float(row["value_usd"])
        groups[key]["total"] += value
        groups[key]["partners"].add(row["partner_iso"])
        groups[key]["codes"].add(row["hts"])
        if row["partner_iso"] == "CHN":
            groups[key]["china"] += value

    result: list[dict[str, Any]] = []
    minerals = sorted({entry[1] for entry in MATERIALS.values()})
    for mineral in minerals:
        for ytd in ("false", "true"):
            for year in YEARS:
                group = groups.get((mineral, year, ytd))
                total = group["total"] if group else 0
                china = group["china"] if group else 0
                result.append(
                    {
                        "mineral": mineral,
                        "year": year,
                        "ytd_flag": ytd,
                        "china_value_usd": round(china, 6) if group else "",
                        "selected_partner_value_usd": round(total, 6) if group else "",
                        "china_share": round(china / total, 10) if total > 0 else "",
                        "reported_partner_count": len(group["partners"]) if group else 0,
                        "hts_codes": "|".join(sorted(group["codes"])) if group else "",
                        "denominator_scope": DENOMINATOR_SCOPE,
                        "data_status": "reported" if total > 0 else "missing_or_zero_denominator",
                    }
                )
    return result


def build_diversification(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, int, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    codes: dict[tuple[str, int, str], set[str]] = defaultdict(set)
    for row in usable_value_rows(rows):
        if row["flow"] != "imports_for_consumption" or row["partner_iso"] == "CHN":
            continue
        key = row["mineral"], int(row["year"]), row["ytd_flag"]
        groups[key][row["partner_iso"]] += float(row["value_usd"])
        codes[key].add(row["hts"])

    result: list[dict[str, Any]] = []
    minerals = sorted({entry[1] for entry in MATERIALS.values()})
    for mineral in minerals:
        for ytd in ("false", "true"):
            for year in YEARS:
                key = mineral, year, ytd
                positive = {iso: value for iso, value in groups.get(key, {}).items() if value > 0}
                total = sum(positive.values())
                hhi = sum((value / total) ** 2 for value in positive.values()) if total > 0 else None
                result.append(
                    {
                        "mineral": mineral,
                        "year": year,
                        "ytd_flag": ytd,
                        "non_china_supplier_count": len(positive),
                        "non_china_value_usd": round(total, 6) if positive else "",
                        "hhi_value_0_1": round(hhi, 10) if hhi is not None else "",
                        "hhi_value_0_10000": round(hhi * 10000, 4) if hhi is not None else "",
                        "hts_codes": "|".join(sorted(codes.get(key, set()))),
                        "denominator_scope": "non_china_members_of_selected_18_partners",
                        "data_status": "reported" if total > 0 else "missing_or_zero_denominator",
                    }
                )
    return result


def unit_basis(qty: int | float | None, unit: str) -> tuple[int | float | None, str]:
    if qty is None:
        return None, unit
    if unit == "metric_ton":
        return qty * 1000, "kg"
    if unit == "kg":
        return qty, "kg"
    return qty, unit


def build_unit_value(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[tuple[str, str, int, str], float] = defaultdict(float)
    eligible: dict[tuple[str, str, int, str, str], dict[str, float]] = defaultdict(
        lambda: {"value": 0.0, "quantity": 0.0}
    )
    for row in usable_value_rows(rows):
        base = row["flow"], row["mineral"], int(row["year"]), row["ytd_flag"]
        totals[base] += float(row["value_usd"])
        if row["quantity_incomplete"] == "true":
            continue
        quantity, basis = unit_basis(row["qty1"], row["qty1_unit"])
        if quantity is None or quantity <= 0 or basis == "not_collected":
            continue
        key = (*base, basis)
        eligible[key]["value"] += float(row["value_usd"])
        eligible[key]["quantity"] += float(quantity)

    result: list[dict[str, Any]] = []
    for key in sorted(eligible):
        flow, mineral, year, ytd, basis = key
        values = eligible[key]
        total = totals[(flow, mineral, year, ytd)]
        quantity = values["quantity"]
        result.append(
            {
                "flow": flow,
                "mineral": mineral,
                "year": year,
                "ytd_flag": ytd,
                "quantity_basis": basis,
                "matched_value_usd": round(values["value"], 6),
                "matched_quantity": round(quantity, 6),
                "unit_value_usd_per_unit": round(values["value"] / quantity, 10),
                "total_value_usd": round(total, 6),
                "value_coverage_share": round(values["value"] / total, 10) if total else "",
                "data_status": "reported_matched_unsuppressed_quantity",
            }
        )
    return result


def rare_earth_disparity(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    values: dict[tuple[int, str, str], float] = defaultdict(float)
    mass: dict[tuple[int, str, str], float] = defaultdict(float)
    incomplete: set[tuple[int, str, str]] = set()
    for row in rows:
        if row["flow"] != "imports_for_consumption" or row["hts"] not in RARE_EARTH_HTS:
            continue
        origin = "china" if row["partner_iso"] == "CHN" else "other_selected"
        key = int(row["year"]), row["ytd_flag"], origin
        if row["quantity_measure_slot"] == "first" and row["value_usd"] is not None:
            values[key] += float(row["value_usd"])
        if row["mass_kg"] is not None:
            mass[key] += float(row["mass_kg"])
        if row["quantity_incomplete"] == "true":
            incomplete.add(key)

    result: list[dict[str, Any]] = []
    for ytd in ("false", "true"):
        for year in YEARS:
            ckey = year, ytd, "china"
            okey = year, ytd, "other_selected"
            cvalue = values.get(ckey)
            ovalue = values.get(okey)
            cmass = mass.get(ckey)
            omass = mass.get(okey)
            result.append(
                {
                    "year": year,
                    "ytd_flag": ytd,
                    "china_value_usd": round(cvalue, 6) if cvalue is not None else None,
                    "other_selected_value_usd": round(ovalue, 6) if ovalue is not None else None,
                    "china_reported_mass_kg": round(cmass, 6) if cmass is not None else None,
                    "other_selected_reported_mass_kg": round(omass, 6) if omass is not None else None,
                    "mass_status": "incomplete_quantity_suppression"
                    if ckey in incomplete or okey in incomplete
                    else "reported_compatible_mass_measures_not_total_tonnage",
                    "hts_codes": "2805|2846|8505",
                    "denominator_scope": DENOMINATOR_SCOPE,
                }
            )
    return result


def selected_supplier_series(rows: list[dict[str, Any]], mineral: str = "rare_earths") -> dict[str, Any]:
    selected = {
        "Australia": "AUS",
        "Myanmar": "MMR",
        "DRC": "COD",
        "Vietnam": "VNM",
        "Estonia": "EST",
        "Canada": "CAN",
    }
    totals: dict[tuple[int, str], float] = defaultdict(float)
    partner_values: dict[tuple[int, str, str], float] = defaultdict(float)
    for row in usable_value_rows(rows):
        if row["flow"] != "imports_for_consumption" or row["mineral"] != mineral:
            continue
        key = int(row["year"]), row["ytd_flag"]
        totals[key] += float(row["value_usd"])
        partner_values[(int(row["year"]), row["ytd_flag"], row["partner_iso"])] += float(row["value_usd"])
    series = []
    for name, iso in selected.items():
        annual = []
        for year in YEARS:
            total = totals.get((year, "false"), 0)
            value = partner_values.get((year, "false", iso), 0)
            annual.append(
                {
                    "year": year,
                    "value_usd": round(value, 6),
                    "share": round(value / total, 10) if total else None,
                }
            )
        series.append({"name": name, "iso3": iso, "annual": annual})
    return {"mineral": mineral, "hts_codes": ["2805", "2846", "8505"], "partners": series}


def load_prc_comtrade() -> dict[str, Any]:
    base = RAW / "un_comtrade_china" / "2846"
    if not base.exists():
        return {
            "status": "not_loaded",
            "message": "China-reporter imports are not present; U.S. exports are not substituted.",
            "source": "UN Comtrade optional input",
        }
    files = sorted(path for path in base.glob("*.json") if re.fullmatch(r"\d{4}\.json", path.name))
    yearly: list[dict[str, Any]] = []
    requested_iso = {"AUS", "MMR", "COD", "VNM", "EST", "CAN"}
    requested: dict[str, list[dict[str, Any]]] = {iso: [] for iso in sorted(requested_iso)}
    for path in files:
        payload = json.loads(path.read_text(encoding="utf-8"))
        records = payload.get("data", payload if isinstance(payload, list) else [])
        if not isinstance(records, list) or not records:
            continue
        period = int(str(records[0].get("period") or path.stem)[:4])
        classification_code = records[0].get("classificationCode")
        partners = []
        world_value = None
        for record in records:
            partner_code = str(record.get("partnerCode", ""))
            partner_iso = str(record.get("partnerISO") or "")
            value = number(record.get("primaryValue"))
            if partner_code == "0" or partner_iso in {"W00", "WLD"}:
                world_value = float(value or 0)
                continue
            if value is None or value <= 0:
                continue
            partners.append(
                {
                    "iso3": partner_iso,
                    "name": record.get("partnerDesc") or partner_iso,
                    "value_usd": float(value),
                }
            )
        partner_total = sum(item["value_usd"] for item in partners)
        denominator = world_value if world_value and world_value > 0 else partner_total
        hhi = sum((item["value_usd"] / partner_total) ** 2 for item in partners) if partner_total else None
        top = sorted(partners, key=lambda item: item["value_usd"], reverse=True)[:5]
        yearly.append(
            {
                "year": period,
                "classification_code": classification_code,
                "trade_system": "Special" if period < 2000 else "General",
                "world_value_usd": world_value,
                "partner_sum_usd": round(partner_total, 6),
                "positive_origin_count": len(partners),
                "hhi_value_0_1": round(hhi, 10) if hhi is not None else None,
                "top_origins": [
                    {
                        **item,
                        "share": round(item["value_usd"] / denominator, 10) if denominator else None,
                    }
                    for item in top
                ],
            }
        )
        by_iso = {item["iso3"]: item for item in partners}
        for iso in requested:
            item = by_iso.get(iso)
            requested[iso].append(
                {
                    "year": period,
                    "value_usd": item["value_usd"] if item else 0,
                    "share": round(item["value_usd"] / denominator, 10) if item and denominator else 0,
                }
            )
    yearly.sort(key=lambda row: row["year"])
    if not yearly:
        return {
            "status": "not_loaded",
            "message": "No valid China-reporter response snapshots were found.",
            "source": "UN Comtrade optional input",
        }
    names = {"AUS": "Australia", "MMR": "Myanmar", "COD": "DRC", "VNM": "Vietnam", "EST": "Estonia", "CAN": "Canada"}
    return {
        "status": "loaded",
        "source": "UN Comtrade, China-reporter annual imports, HS 2846",
        "coverage": [yearly[0]["year"], yearly[-1]["year"]],
        "scope": "All reported origins; CIF import value; original HS edition by year.",
        "causal_caution": "Trade statistics show changing reported origins, not proof of state cultivation.",
        "yearly": yearly,
        "selected_partners": [
            {"name": names[iso], "iso3": iso, "annual": sorted(requested[iso], key=lambda row: row["year"])}
            for iso in ("AUS", "MMR", "COD", "VNM", "EST", "CAN")
        ],
    }


def comtrade_source_metadata() -> dict[str, Any] | None:
    path = RAW / "un_comtrade_china" / "2846" / "manifest.json"
    if not path.is_file():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {
        "flow": "china_reported_imports",
        "file": "data/raw/un_comtrade_china/2846/manifest.json",
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
        "source": "UN Comtrade public preview API",
        "coverage": [payload["summary"]["first_year"], payload["summary"]["last_year"]],
        "raw_response_count": payload["summary"]["year_count"],
        "commodity": "HS 2846",
        "validation_status": payload["validation"]["status"],
    }


def build_explorer(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    index: list[dict[str, Any]] = []
    by_hts: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_hts[row["hts"]].append(row)
    for hts in sorted(by_hts):
        label, mineral, stage, scope_note = MATERIALS[hts]
        source_rows = by_hts[hts]
        descriptions = sorted({row["hts_desc"] for row in source_rows})
        units = sorted({row["qty1_unit"] or row["qty2_unit"] for row in source_rows if row["qty1_unit"] or row["qty2_unit"]})
        groups: dict[tuple[str, str, str, int, str], dict[str, Any]] = defaultdict(
            lambda: {
                "value": 0.0,
                "value_seen": False,
                "mass": 0.0,
                "mass_seen": False,
                "matched_value": 0.0,
                "matched_mass": 0.0,
                "matched_seen": False,
                "incomplete": False,
            }
        )
        for row in source_rows:
            key = row["flow"], row["partner"], row["partner_iso"], int(row["year"]), row["ytd_flag"]
            group = groups[key]
            if row["value_usd"] is not None:
                group["value"] += float(row["value_usd"])
                group["value_seen"] = True
            if row["mass_kg"] is not None:
                group["mass"] += float(row["mass_kg"])
                group["mass_seen"] = True
            if (
                row["quantity_measure_slot"] == "first"
                and row["value_usd"] is not None
                and row["mass_kg"] is not None
                and float(row["mass_kg"]) > 0
                and row["quantity_incomplete"] != "true"
            ):
                group["matched_value"] += float(row["value_usd"])
                group["matched_mass"] += float(row["mass_kg"])
                group["matched_seen"] = True
            group["incomplete"] = group["incomplete"] or row["quantity_incomplete"] == "true"
        data = []
        for key in sorted(groups):
            flow, partner, iso, year, ytd = key
            group = groups[key]
            value = group["value"] if group["value_seen"] else None
            mass = group["mass"] if group["mass_seen"] else None
            matched_value = group["matched_value"] if group["matched_seen"] else None
            matched_mass = group["matched_mass"] if group["matched_seen"] else None
            unit_value = matched_value / matched_mass if matched_value is not None and matched_mass else None
            data.append(
                [
                    flow,
                    partner,
                    iso,
                    year,
                    ytd == "true",
                    round(value, 6) if value is not None else None,
                    round(mass, 6) if mass is not None else None,
                    round(unit_value, 10) if unit_value is not None else None,
                    "quantity_incomplete" if group["incomplete"] else "reported",
                    round(matched_mass, 6) if matched_mass is not None else None,
                ]
            )
        payload = {
            "schema_version": "3.1.0",
            "hts": hts,
            "label": label,
            "descriptions": descriptions,
            "mineral": mineral,
            "processing_stage": stage,
            "scope_note": scope_note,
            "source_units": units,
            "quantity_note": "Reported mass combines compatible kg measures across first and independent second quantity slots. It is coverage, not necessarily total physical tonnage. Unit value uses only value and mass from the same unsuppressed first-quantity bucket.",
            "columns": [
                "flow",
                "partner",
                "partner_iso",
                "year",
                "ytd_flag",
                "value_usd",
                "reported_mass_kg",
                "unit_value_usd_per_reported_kg",
                "measurement_status",
                "unit_value_matched_mass_kg",
            ],
            "rows": data,
        }
        path = EXPLORER / f"{hts}.json"
        path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        index.append(
            {
                "hts": hts,
                "label": label,
                "mineral": mineral,
                "processing_stage": stage,
                "scope_note": scope_note,
                "file": f"data/processed/explorer/{hts}.json",
                "bytes": path.stat().st_size,
            }
        )
    return index


def build_site_summary(
    rows: list[dict[str, Any]],
    shares: list[dict[str, Any]],
    diversification: list[dict[str, Any]],
    explorer_index: list[dict[str, Any]],
    sources: list[dict[str, Any]],
    usgs_rows: list[dict[str, Any]],
    usgs_metadata: dict[str, Any],
) -> dict[str, Any]:
    disparity = rare_earth_disparity(rows)
    latest = next(
        row
        for row in shares
        if row["mineral"] == "rare_earths" and row["year"] == 2025 and row["ytd_flag"] == "false"
    )
    share_series_annual: dict[str, list[dict[str, Any]]] = defaultdict(list)
    share_series_ytd: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in shares:
        if row["ytd_flag"] == "false" and row["year"] <= 2025:
            share_series_annual[row["mineral"]].append(
                {"year": row["year"], "share": row["china_share"] if row["china_share"] != "" else None}
            )
        elif row["ytd_flag"] == "true":
            share_series_ytd[row["mineral"]].append(
                {"year": row["year"], "share": row["china_share"] if row["china_share"] != "" else None}
            )
    rare_div = [
        row
        for row in diversification
        if row["mineral"] == "rare_earths" and row["ytd_flag"] == "false" and row["year"] <= 2025
    ]
    public_sources = [
        {"flow": item["flow"], "file": item["file"], "sha256": item["sha256"], "bytes": item["bytes"]}
        for item in sources
    ]
    comtrade_source = comtrade_source_metadata()
    if comtrade_source:
        public_sources.append(comtrade_source)
    public_sources.append(
        {
            "flow": "usgs_rare_earths_context",
            "file": usgs_metadata["file"],
            "sha256": usgs_metadata["sha256"],
            "bytes": usgs_metadata["bytes"],
            "source": usgs_metadata["title"],
            "coverage": usgs_metadata["coverage"],
        }
    )
    return {
        "schema_version": "3.2.0",
        "generated_at": f"{RETRIEVED_AT}T00:00:00Z",
        "title": "US-PRC Critical Minerals Record, 1993-2026",
        "coverage": {
            "usitc_annual": [1993, 2025],
            "usitc_ytd": [1993, 2026],
            "ytd_months": "January-April",
            "selected_partner_count": 18,
            "us_denominator_scope": DENOMINATOR_SCOPE,
            "usgs_ds140": [1900, 2020],
        },
        "headline": {
            "year": 2025,
            "value": latest["china_share"],
            "label": "China share of U.S. rare-earth-proxy imports among 18 selected origins",
            "numerator_usd": latest["china_value_usd"],
            "denominator_usd": latest["selected_partner_value_usd"],
            "hts_codes": ["2805", "2846", "8505"],
        },
        "rare_earth_share_annual": share_series_annual["rare_earths"],
        "disparity": disparity,
        "china_share_by_mineral": {
            mineral: {
                "annual": share_series_annual[mineral],
                "ytd": share_series_ytd[mineral],
            }
            for mineral in sorted(set(share_series_annual) | set(share_series_ytd))
        },
        "us_non_china_diversification": rare_div,
        "us_selected_supplier_series": selected_supplier_series(rows),
        "prc_supply_origins": load_prc_comtrade(),
        "usgs_rare_earths_context": build_usgs_site_context(usgs_rows, usgs_metadata),
        "explorer_index": explorer_index,
        "sources": public_sources,
        "warnings": [
            "The DataWeb denominator is the sum of 18 selected partners, not World.",
            "HTS 2805 and 8505 are broad proxies, not pure rare-earth categories.",
            "2026 is January-April YTD; no 2026 annual value is plotted.",
            "HTS 8505 has a reported-quantity regime break in 2019.",
            "Second quantities are preserved separately because the HTS4 export provides no lossless value-bucket join.",
        ],
    }


def write_classification_breaks() -> None:
    rows = [
        {
            "break_note_id": f"hs-revision-{year}",
            "effective_year": year,
            "hts_scope": "all_selected_hts4",
            "status": "revision_boundary_continuity_not_assessed",
            "note": f"New U.S. HTS edition aligned with the HS {year} revision; continuity for each four-digit proxy requires a concordance review.",
            "source_url": "https://www.usitc.gov/harmonized_tariff_information/hts/archive/list",
        }
        for year in (1996, 2002, 2007, 2012, 2017, 2022)
    ]
    rows.append(
        {
            "break_note_id": "hts8505-unit-regime-2019",
            "effective_year": 2019,
            "hts_scope": "8505",
            "status": "observed_measurement_break",
            "note": "DataWeb values shift sharply from 'no units collected' to kilogram-denominated buckets between 2018 and 2019; reported tonnage is not directly comparable across this boundary.",
            "source_url": "data/raw/us_imports_for_consumption_1993-2026.xlsx",
        }
    )
    write_csv(
        PROCESSED / "classification_breaks.csv",
        rows,
        ["break_note_id", "effective_year", "hts_scope", "status", "note", "source_url"],
    )


def write_data_dictionary() -> None:
    descriptions = {
        "reporter": "Reporting economy; US for both DataWeb workbooks.",
        "flow": "imports_for_consumption or domestic_exports.",
        "partner": "Country of origin for imports; ultimate destination for exports.",
        "partner_iso": "ISO 3166-1 alpha-3 partner code.",
        "hts": "Four-digit HTS heading as queried.",
        "hts_desc": "Source-reported heading description.",
        "mineral": "Analytical mineral/proxy family.",
        "processing_stage": "ore, metal, compound, magnet, or alloy.",
        "year": "Calendar year.",
        "ytd_flag": "true for comparable January-April YTD; false for full-year annual.",
        "value_usd": "Customs value for imports or FAS value for domestic exports; blank on Q2-only rows.",
        "qty1": "First unit quantity on the losslessly matched value bucket.",
        "qty1_unit": "Normalized first quantity unit.",
        "qty2": "Independent second-unit quantity row; never duplicated across value buckets.",
        "qty2_unit": "Normalized second quantity unit.",
        "source": "Publishing agency/system.",
        "retrieved_at": "Frozen retrieval date.",
    }
    rows = [
        {"column": column, "type": "string_or_number", "description": descriptions.get(column, "ETL provenance or measurement-quality field.")}
        for column in CSV_COLUMNS
    ]
    write_csv(PROCESSED / "data_dictionary.csv", rows, ["column", "type", "description"])


def write_usgs_data_dictionary() -> None:
    descriptions = {
        "source_id": "Stable identifier for this frozen USGS workbook vintage.",
        "commodity": "Normalized commodity family; rare_earths for this source.",
        "year": "Calendar year in the USGS worksheet.",
        "period_type": "Annual period; the source contains no YTD observations.",
        "geography": "United States or World, according to the source heading.",
        "geography_code": "USA or WLD analytical geography code; WLD is not an ISO-3166 code.",
        "metric": "Normalized metric identifier.",
        "metric_label": "Original USGS worksheet column heading.",
        "value": "Numeric source value; blank for NA or W cells.",
        "unit": "REO-equivalent mass or unit-value basis.",
        "value_status": "available, not_available, or withheld.",
        "method_status": "Method explicitly supported by the embedded worksheet notes; no invented reported/estimated distinction.",
        "method_note_id": "Embedded worksheet-notes section supporting the method and scope label.",
        "source_value_raw": "Original displayed cell token or number.",
        "source_formula": "Exact source formula when the workbook cell contains one; otherwise blank.",
        "scope_note": "Metric definition distilled from the workbook's embedded notes document.",
        "source": "Publishing agency and data-series label.",
        "source_file": "Frozen repository path.",
        "source_sheet": "Source worksheet name.",
        "source_cell": "Exact source cell address.",
        "source_url": "Official USGS landing page.",
        "download_url": "Official source-file URL.",
        "worksheet_last_modified": "Last-modification date printed in the worksheet.",
        "package_modified_at": "Modification timestamp in the XLSX core properties.",
        "source_publication_date": "Publication date on the USGS media page.",
        "retrieved_at": "Repository retrieval date.",
    }
    rows = [
        {"column": column, "type": "string_or_number", "description": descriptions[column]}
        for column in USGS_DS140_COLUMNS
    ]
    write_csv(
        PROCESSED / "usgs_rare_earths_data_dictionary.csv",
        rows,
        ["column", "type", "description"],
    )


def write_prc_origin_index(prc: dict[str, Any]) -> int:
    if prc.get("status") != "loaded":
        return 0
    rows: list[dict[str, Any]] = []
    for row in prc["yearly"]:
        leader = row["top_origins"][0] if row["top_origins"] else {}
        rows.append(
            {
                "reporter": "China",
                "flow": "imports",
                "hts": "2846",
                "year": row["year"],
                "classification_code": row["classification_code"],
                "trade_system": row["trade_system"],
                "world_value_usd": row["world_value_usd"],
                "partner_sum_usd": row["partner_sum_usd"],
                "positive_origin_count": row["positive_origin_count"],
                "hhi_value_0_1": row["hhi_value_0_1"],
                "leading_origin": leader.get("name", ""),
                "leading_origin_iso": leader.get("iso3", ""),
                "leading_origin_value_usd": leader.get("value_usd", ""),
                "leading_origin_share": leader.get("share", ""),
                "source": "UN Comtrade",
                "retrieved_at": RETRIEVED_AT,
                "data_status": "reported_all_origins_reconciled_to_world",
            }
        )
    write_csv(
        PROCESSED / "prc_supplier_origin_index.csv",
        rows,
        [
            "reporter", "flow", "hts", "year", "classification_code", "trade_system",
            "world_value_usd", "partner_sum_usd", "positive_origin_count", "hhi_value_0_1",
            "leading_origin", "leading_origin_iso", "leading_origin_value_usd",
            "leading_origin_share", "source", "retrieved_at", "data_status",
        ],
    )
    return len(rows)


def main() -> int:
    global PROCESSED, EXPLORER
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=PROCESSED, help="Processed-data directory")
    args = parser.parse_args()
    PROCESSED = args.output.resolve()
    EXPLORER = PROCESSED / "explorer"
    if PROCESSED.exists():
        shutil.rmtree(PROCESSED)
    EXPLORER.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []
    for config in FLOWS:
        flow_rows, metadata = emit_flow(config)
        rows.extend(flow_rows)
        sources.append(metadata)
    rows.sort(
        key=lambda row: (
            row["flow"],
            row["partner_iso"],
            row["hts"],
            row["ytd_flag"],
            row["quantity_measure_slot"],
            row["source_quantity_bucket"],
            int(row["year"]),
        )
    )
    write_csv(PROCESSED / "trade_long.csv", rows, CSV_COLUMNS)

    usgs_rows, usgs_metadata = parse_usgs_ds140()
    write_csv(
        PROCESSED / "usgs_rare_earths_historical.csv",
        usgs_rows,
        USGS_DS140_COLUMNS,
    )
    (PROCESSED / "usgs_rare_earths_metadata.json").write_text(
        json.dumps(usgs_metadata, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    shares = build_china_share(rows)
    share_columns = [
        "mineral", "year", "ytd_flag", "china_value_usd", "selected_partner_value_usd",
        "china_share", "reported_partner_count", "hts_codes", "denominator_scope", "data_status",
    ]
    write_csv(PROCESSED / "china_share_of_us_imports.csv", shares, share_columns)

    diversification = build_diversification(rows)
    diversification_columns = [
        "mineral", "year", "ytd_flag", "non_china_supplier_count", "non_china_value_usd",
        "hhi_value_0_1", "hhi_value_0_10000", "hts_codes", "denominator_scope", "data_status",
    ]
    write_csv(PROCESSED / "supplier_diversification_index.csv", diversification, diversification_columns)

    unit_values = build_unit_value(rows)
    write_csv(
        PROCESSED / "unit_value.csv",
        unit_values,
        [
            "flow", "mineral", "year", "ytd_flag", "quantity_basis", "matched_value_usd",
            "matched_quantity", "unit_value_usd_per_unit", "total_value_usd", "value_coverage_share", "data_status",
        ],
    )

    write_classification_breaks()
    write_data_dictionary()
    write_usgs_data_dictionary()
    explorer_index = build_explorer(rows)

    comtrade_source = comtrade_source_metadata()
    manifest = {
        "schema_version": "3.2.0",
        "generated_at": f"{RETRIEVED_AT}T00:00:00Z",
        "source_system": "USITC DataWeb",
        "source_url": SOURCE_URL,
        "reporter": "United States",
        "retrieved_at": RETRIEVED_AT,
        "classification": "HTS Items",
        "aggregation_level": 4,
        "annual_coverage": [1993, 2025],
        "ytd_coverage": [1993, 2026],
        "ytd_months": "January-April",
        "denominator_scope": DENOMINATOR_SCOPE,
        "denominator_warning": "The workbooks contain 18 selected partners and no World row.",
        "value_quantity_join": "Value and first quantity join exactly by partner, HTS4, source description, and normalized source quantity bucket.",
        "second_quantity_rule": "Second quantities remain independent rows because the HTS4 export has no lossless key to a first-unit value bucket.",
        "sources": sources,
        "optional_china_reporter_source": comtrade_source,
        "usgs_rare_earths_source": usgs_metadata,
        "official_method_links": {
            "partner_definitions": "https://www.usitc.gov/faq/question/what_meant_country_merchandise_trade_statistics.htm",
            "dataweb_api": "https://www.usitc.gov/applications/dataweb/api/dataweb_query_api.html",
            "dataweb_quantity_faq": "https://www.usitc.gov/applications/dataweb/faqs",
            "hts_archive": "https://www.usitc.gov/harmonized_tariff_information/hts/archive/list",
            "usgs_ds140_rare_earths": USGS_DS140_PAGE_URL,
        },
        "processed": {
            "trade_long_rows": len(rows),
            "china_share_rows": len(shares),
            "diversification_rows": len(diversification),
            "unit_value_rows": len(unit_values),
            "explorer_files": len(explorer_index),
            "usgs_rare_earths_rows": len(usgs_rows),
        },
    }
    (PROCESSED / "explorer-index.json").write_text(
        json.dumps(explorer_index, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    summary = build_site_summary(
        rows,
        shares,
        diversification,
        explorer_index,
        sources,
        usgs_rows,
        usgs_metadata,
    )
    prc_origin_rows = write_prc_origin_index(summary["prc_supply_origins"])
    manifest["processed"]["prc_supplier_origin_index_rows"] = prc_origin_rows
    (PROCESSED / "query_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    (PROCESSED / "site-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8"
    )

    print(
        f"Built {len(rows):,} DataWeb rows, {len(usgs_rows):,} USGS DS140 rows, "
        f"{len(unit_values):,} unit-value rows, and {len(explorer_index)} explorer shards "
        f"in {PROCESSED.relative_to(ROOT)}."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
