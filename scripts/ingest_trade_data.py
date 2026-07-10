#!/usr/bin/env python3
"""Build verified U.S. strategic-resource trade context for 1861-1992.

Annual commodity imports and exports come from official USGS Data Series 140
workbooks. The 1861-1900 bridge uses the Census Bureau's published multi-year
averages for the broad economic class "crude materials." The two series remain
separate and are never combined, interpolated, or treated as bilateral trade.
"""

from __future__ import annotations

import argparse
import json
import math
import tempfile
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from ingest_usgs_ds140 import (
    BASE,
    ROOT,
    SOURCES,
    download,
    normalized_header,
    unit_for,
    workbook_unit_text,
)


DEFAULT_OUTPUT = ROOT / "data" / "history-stack" / "trade.json"
CENSUS_SOURCE_URL = "https://www2.census.gov/library/publications/1948/compendia/statab/69ed/1948-12.pdf"

# Values are published yearly averages in thousands of current dollars.
# Shares are the corresponding published percent distributions in table 1014.
CENSUS_PERIODS = [
    (1861, 1865, 170198, 33990, 19.97, 255439, 36064, 14.12),
    (1866, 1870, 307696, 177296, 57.62, 408295, 47814, 11.71),
    (1871, 1875, 486128, 218449, 44.94, 577873, 93182, 16.12),
    (1876, 1880, 663650, 213989, 32.24, 492570, 91353, 18.55),
    (1881, 1885, 774607, 261645, 33.78, 667142, 133268, 19.98),
    (1886, 1890, 725685, 276703, 38.13, 717231, 162436, 22.65),
    (1891, 1895, 876326, 295087, 33.67, 785137, 185222, 23.59),
    (1896, 1900, 1136039, 296684, 26.11, 741519, 218517, 29.47),
]


def source_page(mineral_id: str) -> str:
    if mineral_id == "bauxite":
        slug = "bauxite-and-alumina"
    elif mineral_id == "rare-earth-elements":
        slug = "rare-earths"
    else:
        slug = mineral_id
    return f"https://www.usgs.gov/media/files/{slug}-historical-statistics-data-series-140"


def extract_usgs_trade(mineral_id: str, cache_dir: Path, access_date: str) -> list[dict]:
    title, filename, publication_year = SOURCES[mineral_id]
    download_url = BASE + filename
    local_path = cache_dir / filename
    if not local_path.exists():
        download(download_url, local_path)

    workbook = load_workbook(local_path, read_only=True, data_only=True)
    worksheet = workbook["Bauxite"] if mineral_id == "bauxite" else workbook.worksheets[0]
    headers = [normalized_header(cell.value) for cell in worksheet[5]]
    workbook_unit = workbook_unit_text(worksheet.cell(3, 1).value)
    rows: list[dict] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=6, values_only=True), start=6):
        year = values[0]
        if not isinstance(year, int) or not 1861 <= year <= 1992:
            continue
        directions = [(column, direction) for column, direction in enumerate(headers) if direction in {"imports", "exports"}]
        for column, direction in directions:
            if column >= len(values):
                continue
            value = values[column]
            if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
                continue
            unit, _ = unit_for(headers[column], workbook_unit)
            rows.append({
                "id": f"usgs-ds140-trade-{mineral_id}-{year}-{direction}",
                "year_start": year,
                "year_end": year,
                "year_label": str(year),
                "temporal_precision": "annual",
                "direction": direction,
                "metric": f"U.S. {direction}",
                "material_scope": "commodity",
                "mineral_id": mineral_id,
                "resource_group": None,
                "partner_scope": "World aggregate; partner countries are not identified in this row.",
                "value": value,
                "unit": unit,
                "trade_basis": f"USGS worksheet header: {worksheet.cell(5, column + 1).value}. Consult the embedded commodity notes for the commodity-specific definition.",
                "calendar_basis": "calendar year unless the commodity workbook notes state otherwise",
                "agency": "U.S. Geological Survey",
                "publication_title": f"{title} historical statistics, Data Series 140",
                "publication_year": publication_year,
                "table_or_page": f"{worksheet.title} worksheet, row {row_number}, column {column + 1} ({worksheet.cell(5, column + 1).value})",
                "source_id": "usgs-ds140",
                "source_url": source_page(mineral_id),
                "download_url": download_url,
                "access_date": access_date,
                "transcription_status": "machine-extracted-xlsx",
                "original_unit": unit,
                "displayed_unit": unit,
                "conversion_methodology": "No project conversion. Value reproduced from the USGS-standardized XLSX cell.",
                "notes": "A blank, withheld, or nonnumeric source cell is omitted rather than treated as zero. This is national aggregate trade, not bilateral trade.",
                "confidence": "high",
            })
    return rows


def census_row(
    start: int,
    end: int,
    direction: str,
    measure: str,
    value: float,
    table: str,
    page: int,
    access_date: str,
) -> dict:
    metric = f"U.S. crude-material {direction} {measure}"
    unit = "percent of total merchandise value" if measure == "share" else "thousand current U.S. dollars"
    return {
        "id": f"census-statab-1948-crude-materials-{start}-{end}-{direction}-{measure}",
        "year_start": start,
        "year_end": end,
        "year_label": f"{start}-{end} yearly average",
        "temporal_precision": "published-multi-year-average",
        "direction": direction,
        "metric": metric,
        "material_scope": "broad-economic-class",
        "mineral_id": None,
        "resource_group": "Crude materials",
        "partner_scope": "All U.S. merchandise trade; partner countries are not identified.",
        "value": value,
        "unit": unit,
        "trade_basis": "Exports of U.S. merchandise" if direction == "exports" else "General imports",
        "calendar_basis": "Published fiscal-year yearly average",
        "agency": "U.S. Department of Commerce, Bureau of the Census",
        "publication_title": "Statistical Abstract of the United States: 1948",
        "publication_year": 1948,
        "table_or_page": f"Table {table}, p. {page}",
        "source_id": "census-statistical-abstract-1948",
        "source_url": CENSUS_SOURCE_URL,
        "download_url": CENSUS_SOURCE_URL,
        "access_date": access_date,
        "transcription_status": "manually-reviewed-published-table",
        "original_unit": unit,
        "displayed_unit": unit,
        "conversion_methodology": "No project conversion or interpolation. The published multi-year average applies to each selected year in the stated period.",
        "notes": "Crude materials is a broad historical economic class that includes minerals and non-mineral raw materials. It is not a mineral-specific or bilateral series.",
        "confidence": "high",
    }


def census_trade(access_date: str) -> list[dict]:
    rows: list[dict] = []
    for start, end, export_total, export_crude, export_share, import_total, import_crude, import_share in CENSUS_PERIODS:
        rows.extend([
            census_row(start, end, "exports", "value", export_crude, "1013", 908, access_date),
            census_row(start, end, "exports", "share", export_share, "1014", 910, access_date),
            census_row(start, end, "imports", "value", import_crude, "1013", 909, access_date),
            census_row(start, end, "imports", "share", import_share, "1014", 910, access_date),
        ])
        # Retain the published denominators in notes without creating derived data.
        rows[-4]["published_total_merchandise_value"] = export_total
        rows[-4]["published_total_merchandise_unit"] = "thousand current U.S. dollars"
        rows[-2]["published_total_merchandise_value"] = import_total
        rows[-2]["published_total_merchandise_unit"] = "thousand current U.S. dollars"
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--cache-dir", type=Path)
    parser.add_argument("--access-date", default=date.today().isoformat())
    args = parser.parse_args()

    if args.cache_dir:
        args.cache_dir.mkdir(parents=True, exist_ok=True)
        rows = [row for mineral_id in SOURCES for row in extract_usgs_trade(mineral_id, args.cache_dir, args.access_date)]
    else:
        with tempfile.TemporaryDirectory(prefix="usgs-trade-") as directory:
            rows = [row for mineral_id in SOURCES for row in extract_usgs_trade(mineral_id, Path(directory), args.access_date)]
    rows.extend(census_trade(args.access_date))
    rows.sort(key=lambda row: (row["year_start"], row["year_end"], row["material_scope"], row.get("mineral_id") or "", row["direction"], row["metric"]))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} trade records to {args.output}")


if __name__ == "__main__":
    main()
