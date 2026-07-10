#!/usr/bin/env python3
"""Extract a bounded 1861-1992 pilot from official USGS Data Series 140 XLSX files.

The script performs no interpolation and writes only numeric cells. NARA data is
not involved. Install ``openpyxl`` from requirements.txt before running.
"""

from __future__ import annotations

import argparse
import json
import math
import tempfile
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = ROOT / "data" / "history-stack" / "statistics.json"
YEARS = {1900, 1910, 1917, 1920, 1930, 1939, 1941, 1942, 1945, 1950, 1953, 1960, 1967, 1970, 1973, 1975, 1980, 1990, 1992}
BASE = "https://d9-wret.s3.us-west-2.amazonaws.com/assets/palladium/production/s3fs-public/media/files/"

SOURCES = {
    "aluminum": ("Aluminum", "ds140-aluminum-2021.xlsx", 2021),
    "bauxite": ("Bauxite and Alumina", "ds140-bauxite-alumina-2021.xlsx", 2021),
    "chromium": ("Chromium", "ds140-chromium-2022.xlsx", 2022),
    "cobalt": ("Cobalt", "ds140-cobalt-2021.xlsx", 2021),
    "copper": ("Copper", "ds140-copper-2020.xlsx", 2020),
    "manganese": ("Manganese", "ds140-manganese-2022.xlsx", 2022),
    "rare-earth-elements": ("Rare Earths", "ds140-rare-earths-2020.xlsx", 2020),
    "tin": ("Tin", "ds140-tin-2021.xlsx", 2021),
    "tungsten": ("Tungsten", "ds140-tungsten-2019.xlsx", 2019),
}

METRICS = {
    "primary production": "U.S. primary production",
    "mine production": "U.S. mine production",
    "production": "U.S. production",
    "imports": "U.S. imports",
    "exports": "U.S. exports",
    "government stocks": "U.S. Government stocks",
    "stocks": "U.S. stocks",
    "apparent consumption": "U.S. apparent consumption",
    "unit value ($/t)": "Unit value",
    "unit value (98$/t)": "Real unit value",
    "world production": "World production",
    "world mine production": "World mine production",
}


def normalized_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def slug(value: str) -> str:
    return "-".join("".join(ch if ch.isalnum() else " " for ch in value.lower()).split())


def unit_for(header: str, workbook_unit: str) -> tuple[str, str]:
    if header == "unit value ($/t)":
        return ("current U.S. dollars per metric ton", "nominal")
    if header == "unit value (98$/t)":
        return ("1998 U.S. dollars per metric ton", "real-1998-dollars")
    return (workbook_unit, "not-price")


def workbook_unit_text(raw: object) -> str:
    text = str(raw or "metric tons")
    text = text.strip("[]")
    if "unless otherwise noted" in text:
        text = text.split("unless otherwise noted", 1)[0].rstrip(" ,")
    text = text.replace("All values are in ", "").replace("All values in ", "")
    return text


def download(url: str, destination: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": "critical-minerals-history-pilot/2.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        destination.write_bytes(response.read())


def extract_file(mineral_id: str, cache_dir: Path, access_date: str) -> list[dict]:
    title, filename, publication_year = SOURCES[mineral_id]
    download_url = BASE + filename
    local_path = cache_dir / filename
    if not local_path.exists():
        download(download_url, local_path)

    workbook = load_workbook(local_path, read_only=True, data_only=True)
    worksheet = workbook["Bauxite"] if mineral_id == "bauxite" else workbook.worksheets[0]
    headers = [normalized_header(cell.value) for cell in worksheet[5]]
    workbook_unit = workbook_unit_text(worksheet.cell(3, 1).value)
    source_url = f"https://www.usgs.gov/media/files/{mineral_id.replace('-', '-')}-historical-statistics-data-series-140"
    if mineral_id == "bauxite":
        source_url = "https://www.usgs.gov/media/files/bauxite-and-alumina-historical-statistics-data-series-140"
    elif mineral_id == "rare-earth-elements":
        source_url = "https://www.usgs.gov/media/files/rare-earths-historical-statistics-data-series-140"

    rows: list[dict] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=6, values_only=True), start=6):
        year = values[0]
        if not isinstance(year, int) or year not in YEARS or not 1861 <= year <= 1992:
            continue
        for column, header in enumerate(headers[1:], start=1):
            if header not in METRICS or column >= len(values):
                continue
            value = values[column]
            if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
                continue
            unit, price_basis = unit_for(header, workbook_unit)
            metric = METRICS[header]
            rows.append({
                "id": f"usgs-ds140-{mineral_id}-{year}-{slug(metric)}",
                "metric": metric,
                "mineral_id": mineral_id,
                "country_id": "united-states" if not metric.startswith("World") else None,
                "year": year,
                "unit": unit,
                "value": value,
                "price_basis": price_basis,
                "publication_title": f"{title} historical statistics, Data Series 140",
                "publication_year": publication_year,
                "table_or_page": f"{worksheet.title} worksheet, row {row_number}, column {column + 1} ({worksheet.cell(5, column + 1).value})",
                "agency": "U.S. Geological Survey",
                "source_url": source_url,
                "download_url": download_url,
                "access_date": access_date,
                "transcription_status": "machine-extracted-xlsx",
                "original_unit": unit,
                "displayed_unit": unit,
                "conversion_methodology": "No project conversion. Value reproduced from the USGS-standardized XLSX cell.",
                "notes": "Data Series 140 may incorporate conversions made by USGS from historical source units. Missing, withheld, and nonnumeric cells are omitted rather than treated as zero.",
                "confidence": "high"
            })
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--cache-dir", type=Path)
    parser.add_argument("--access-date", default=date.today().isoformat())
    args = parser.parse_args()

    if args.cache_dir:
        args.cache_dir.mkdir(parents=True, exist_ok=True)
        rows = [row for mineral_id in SOURCES for row in extract_file(mineral_id, args.cache_dir, args.access_date)]
    else:
        with tempfile.TemporaryDirectory(prefix="usgs-ds140-") as directory:
            rows = [row for mineral_id in SOURCES for row in extract_file(mineral_id, Path(directory), args.access_date)]
    rows.sort(key=lambda row: (row["mineral_id"], row["year"], row["metric"]))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} statistics to {args.output}")


if __name__ == "__main__":
    main()
